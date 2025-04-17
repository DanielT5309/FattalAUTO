import traceback
from unittest import TestCase
from selenium import webdriver
import random
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Fattal_Pages.Fattal_Main_Page import FattalMainPage
from Fattal_Pages.Fattal_Tool_Bar import FattalToolBar
from Fattal_Pages.Fattal_Search_Result import FattalSearchResultPage
from Fattal_Pages.Fattal_Order_Page_ISR import FattalOrderPage
from Fattal_Pages.Fattal_Flight_OrderPage import FattalFlightOrderPage
import logging
import platform
from datetime import datetime
from faker import Faker
import os
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import io
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys

class FattalTests(TestCase):
    def setUp(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--force-device-scale-factor=0.75")

        self.test_start_time = datetime.now()
        self.log_stream = io.StringIO()

        # Calculate base dir to Fattal_Tests (where test_fattal.py lives)
        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(self.log_stream),
                logging.StreamHandler(sys.stdout)
            ]
        )

        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")

        self.driver = webdriver.Chrome(options=options)
        #self.driver.get("https://www.fattal.co.il")
        self.driver.get("https://fe.stage.fattal.kimaia.dev/")
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)

        self.main_page = FattalMainPage(self.driver)
        self.toolbar = FattalToolBar(self.driver)
        self.search_result = FattalSearchResultPage(self.driver)
        self.order_page = FattalOrderPage(self.driver)
        self.flight_page = FattalFlightOrderPage(self.driver)
        self.driver.execute_script("window.performance.mark('selenium-start')")
        self.addCleanup(lambda: self.post_test_logging(getattr(self, '_outcome', None)))

    def retry_flight_search_if_no_results(self, hotel_name):
        if self.search_result.no_results_found():
            logging.warning("No flight hotel results found. Retrying flight search...")
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.select_flight_option_all_airports()
            self.main_page.search_button()
        else:
            logging.info("Flight search results returned.")

    def post_test_logging(self, result):
        test_method = self._testMethodName
        has_failed = False
        error_msg = ""
        screenshot_path = ""
        log_file_path = ""
        duration = (datetime.now() - self.test_start_time).total_seconds()

        try:
            browser = self.driver.capabilities['browserName'] if self.driver else "unknown"
        except Exception as e:
            logging.warning(f"Could not get browser info: {e}")
            browser = "unknown"

        os_name = platform.system()

        #  Save logs
        log_summary = self.log_stream.getvalue()
        logs_dir = os.path.join(self.base_dir, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        log_file_path = os.path.join(logs_dir, f"{test_method}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        try:
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(log_summary)
            print(f"[LOG DEBUG] Written log to {log_file_path}")
        except Exception as e:
            print(f"[LOG DEBUG] Failed to write log: {e}")

        #  Robust failure detection that works in PyCharm + unittest
        try:
            outcome = getattr(result, 'result', result)
            for failed_test, exc_info in outcome.failures + outcome.errors:
                if self._testMethodName in str(failed_test):
                    has_failed = True
                    if len(exc_info) == 3:
                        exc_type, exc_value, tb = exc_info
                        error_msg = "".join(traceback.format_exception(exc_type, exc_value, tb))
                    else:
                        error_msg = f" Unrecognized exception format: {exc_info}"
                    break
        except Exception as e:
            logging.warning(f" Failed to analyze test outcome: {e}")

        # 📸 Screenshot if test failed
        if has_failed:
            try:
                if self.driver:
                    screenshot_dir = os.path.join(self.base_dir, "Screenshots")
                    os.makedirs(screenshot_dir, exist_ok=True)
                    screenshot_path = os.path.join(
                        screenshot_dir, f"{test_method}_FAIL_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
                    )
                    self.driver.save_screenshot(screenshot_path)
                    logging.error(f" Screenshot saved at: {screenshot_path}")
            except Exception as e:
                screenshot_path = ""
                logging.warning(f" Could not save screenshot: {e}")

        #  Build test info dict
        test_info = {
            "name": test_method,
            "description": "Test Description",
            "status": "FAILED" if has_failed else "PASSED",
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "duration": f"{duration:.2f}s",
            "browser": browser,
            "os": os_name,
            "error": error_msg,
            "screenshot": screenshot_path,
            "log": log_file_path
        }

        self.save_to_excel(test_info)
        self.save_to_pdf(test_info)

    def save_to_excel(self, info: dict):
        parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(parent_folder, "TestResults.xlsx")

        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results"
                ws.append([
                    "Test Name", "Description", "Status", "Timestamp", "Duration",
                    "Browser", "OS", "URL", "Screenshot", "Log File"
                ])
            else:
                wb = load_workbook(filename)
                ws = wb.active

            row = [
                info.get("name", ""),
                info.get("description", ""),
                info.get("status", ""),
                info.get("timestamp", ""),
                info.get("duration", ""),
                info.get("browser", ""),
                info.get("os", ""),
                info.get("url", ""),
                info.get("screenshot", ""),
                info.get("log", "")
            ]
            ws.append(row)
            row_num = ws.max_row

            # Add hyperlink to screenshot
            screenshot_path = info.get("screenshot", "")
            if screenshot_path:
                cell = ws.cell(row=row_num, column=9)
                cell.hyperlink = screenshot_path
                cell.font = Font(color="0000EE", underline="single")

            # Add hyperlink to log file
            log_path = info.get("log", "")
            if log_path:
                cell = ws.cell(row=row_num, column=10)
                cell.hyperlink = f'file:///{log_path}'
                cell.font = Font(color="0000EE", underline="single")

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.error(f"Error adjusting column width: {e}")
                ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(filename)
            logging.info(f"✔️ Excel file saved at: {filename}")

        except PermissionError as e:
            logging.warning(f"Excel file is open or locked: {e}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")

    def save_to_pdf(self, info: dict):
        pdf_dir = "reports"
        os.makedirs(pdf_dir, exist_ok=True)
        filename = f"{pdf_dir}/{info['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        c = canvas.Canvas(filename, pagesize=A4)

        width, height = A4
        y = height - 72

        c.setFont("Helvetica-Bold", 18)
        c.drawString(72, y, "Test Execution Report")
        y -= 40

        c.setFont("Helvetica", 12)
        lines = [
            f"Test Name: {info.get('name')}",
            f"Description: {info.get('description')}",
            f"Status: {info.get('status')}",
            f"Timestamp: {info.get('timestamp')}",
            f"Browser: {info.get('browser')}",
            f"OS: {info.get('os')}",
        ]
        if info.get("error"):
            lines.append(f"Error: {info.get('error')}")

        for line in lines:
            c.drawString(72, y, line)
            y -= 20

        if info.get("screenshot") and os.path.exists(info["screenshot"]):
            c.drawString(72, y, "Screenshot:")
            y -= 250
            c.drawImage(info["screenshot"], 72, y, width=5 * inch, preserveAspectRatio=True, mask='auto')

        c.showPage()
        c.save()
        print(f" PDF saved: {filename}")


    def run(self, result=None):
        self._resultForDoCleanups = result
        return super().run(result)

    def take_screenshot(self, test_method_name):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        screenshot_dir = os.path.join(os.getcwd(), "Fattal_Tests", "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = f"{screenshot_dir}/{test_method_name}_{timestamp}.png"
        self.driver.save_screenshot(filename)
        logging.error(f" Screenshot taken: {filename}")

    def retry_search_if_no_results(self, hotel_name, adults, children, infants):
        button_text_options = ["להזמנת חדר", "הצג מחירים"]
        buttons = None
        for text in button_text_options:
            try:
                # Wait up to 30 seconds for multiple hotel cards to load
                buttons = WebDriverWait(self.driver, 60).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//button[normalize-space(text())='{text}']"))
                )
                logging.info("Search results returned — continuing.")
            except TimeoutException:
                logging.warning(" No hotels matched the search after waiting.")
                logging.warning(" No hotel results found — restarting search...")
                self.toolbar.logo_mainpage()

                # Allow short pause for main page to stabilize
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.ID, "main-input"))
                )

                self.main_page.click_clear_button_hotel()
                self.main_page.set_city(hotel_name)
                self.main_page.select_next_month_date_range()
                self.main_page.search_button()
                logging.info("Search restarted after empty result.")

    def complete_booking_post_flight(self):
        # Ensure we are NOT in iframe
        self.order_page.switch_to_default_content()

        # Wait until order form is interactable
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.ID, "checkout.personal_details_form.label_email"))
        )
        logging.info(" Order form is ready for input.")

        #  Fill form details
        self.order_page.set_email("test@example.com")
        self.order_page.set_phone("0501234567")
        self.order_page.set_first_name("דניאל")
        self.order_page.set_last_name("טסט")
        self.order_page.set_id_number("315635250")

        #  Accept terms to unlock form
        self.order_page.click_terms_approval_checkbox_js()

        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("טקסט לדוגמא .....")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        #  Payment stage
        self.order_page.switch_to_payment_iframe()

        self.order_page.set_card_number("4580080111866879")
        self.order_page.select_expiry_month("08")
        self.order_page.select_expiry_year("2025")
        self.order_page.set_cvv("955")
        self.order_page.set_cardholder_name("Israel Cohen")
        self.order_page.set_id_number_card("0356998")

        self.order_page.switch_to_default_content()
        logging.info("Booking flow completed after flight selection")

    def complete_booking_flow(self, hotel_name, adults, children, infants):
        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f"Guests selected: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()

        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.ID, "checkout.personal_details_form.label_email"))
        )
        logging.info(" Order form is ready for input.")

        #  Store values for dynamic assertion later
        self.entered_email = "test@example.com"
        self.entered_phone = "0501234567"
        self.entered_first_name = "דניאל"
        self.entered_last_name = "טסט"
        self.entered_id = "315635250"

        self.order_page.set_email(self.entered_email)
        self.order_page.set_phone(self.entered_phone)
        self.order_page.set_first_name(self.entered_first_name)
        self.order_page.set_last_name(self.entered_last_name)
        self.order_page.set_id_number(self.entered_id)

        self.order_page.click_terms_approval_checkbox_js()

        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("טקסט לדוגמא .....")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number("4580080111866879")
        self.order_page.select_expiry_month("08")
        self.order_page.select_expiry_year("2025")
        self.order_page.set_cvv("955")
        self.order_page.set_cardholder_name("Israel Cohen")
        self.order_page.set_id_number_card("0356998")
        self.order_page.switch_to_default_content()

        logging.info(" Booking flow completed (test mode)")

    def complete_booking_flow_club_checkbox(self, hotel_name, adults, children, infants):
        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f"Guests selected: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results( hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()

        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.ID, "checkout.personal_details_form.label_email"))
        )
        logging.info(" Order form is ready for input.")

        self.order_page.club_checkbox()

        # Store values for dynamic assertion later
        self.entered_email = "test@example.com"
        self.entered_phone = "0501234567"
        self.entered_first_name = "דניאל"
        self.entered_last_name = "טסט"
        self.entered_id = "315635250"

        self.order_page.set_email(self.entered_email)
        self.order_page.set_phone(self.entered_phone)
        self.order_page.set_first_name(self.entered_first_name)
        self.order_page.set_last_name(self.entered_last_name)
        self.order_page.set_id_number(self.entered_id)

        self.order_page.click_terms_approval_checkbox_js()

        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("טקסט לדוגמא .....")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number("4580080111866879")
        self.order_page.select_expiry_month("08")
        self.order_page.select_expiry_year("2025")
        self.order_page.set_cvv("955")
        self.order_page.set_cardholder_name("Israel Cohen")
        self.order_page.set_id_number_card("0356998")
        self.order_page.switch_to_default_content()

        logging.info("Booking flow completed (test mode)")

    def test_anonymous_no_login(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        self.complete_booking_flow(hotel_name, adults, children, infants)
        #DO NOT Click the Order Button!!!
        #self.order_page.get_submit_button()
    def test_anonymous_no_login_club_checkbox(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 0, 0

        self.complete_booking_flow_club_checkbox(hotel_name, adults, children, infants)
        #DO NOT Click the Order Button!!!
        #self.order_page.get_submit_button()

    def test_eilat_flight(self):
        hotel_name = "אילת,ישראל"
        adults, children, infants = 3, 1, 1

        try:
            logging.info(" Starting test for Eilat zone with flight")

            # --- HOTEL + ROOM + FLIGHT SELECTION ---
            self.main_page.deal_popup()
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.select_flight_option_all_airports()

            self.main_page.set_room_occupants(adults, children, infants)
            self.main_page.search_button()

            self.retry_flight_search_if_no_results(hotel_name)
            # self.search_result.wait_for_rooms_to_load()
            self.search_result.click_book_room_button()
            self.search_result.wait_for_prices_to_load()
            self.search_result.click_first_show_prices()
            self.search_result.click_first_book_room()

            # --- FLIGHT SELECTION LOGIC ---
            try:
                self.flight_page.try_flight_options_by_time_of_day()
            except Exception as e:
                logging.warning(f" Flight selection failed (non-blocking): {e}")

            # --- PASSENGER FORM ---
            handled = self.flight_page.handle_passenger_form_if_flight_selection_skipped()
            if not handled:
                self.flight_page.wait_for_passenger_form()
                # self.flight_page.fill_passenger_details_and_validate()
                self.flight_page.click_continue_button()

            #  Booking flow continues only if everything succeeded
            self.complete_booking_post_flight()
        except Exception as e:
            logging.error(f" Test failed: {e}")
            try:
                self.take_screenshot("test_eilat_flight")
            except Exception:
                logging.warning("📸 Couldn't capture screenshot due to earlier failure.")
            raise

    def test_eilat(self):
        hotel_name = "אילת,ישראל"
        logging.info(" Starting test: hotel search and booking flow")
        adults, children, infants = 2, 0, 0
        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()
        self.main_page.set_room_occupants(adults, children, infants)
        self.main_page.search_button()
        self.search_result.wait_for_rooms_to_load()
        self.search_result.click_book_room_button()
        self.search_result.wait_for_prices_to_load()
        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()
        self.complete_booking_post_flight()

    def test_club_login(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.deal_popup()

        try:
            self.toolbar.personal_zone()
            WebDriverWait(self.driver, 5).until(
                EC.visibility_of(self.toolbar.user_id_input())
            ).send_keys("999318330")

            self.toolbar.user_password_input().send_keys("Aa123456")
            self.toolbar.login_button()
            self.main_page.close_post_login_popup()

            logging.info("Logged in to club account successfully.")
        except Exception as e:
            logging.warning(f"Login failed: {e}")

        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        self.complete_booking_flow(hotel_name, adults, children, infants)
        # DO NOT Click the Order Button!!!
        # self.order_page.get_submit_button()

    def test_full_random_no_login(self):
        fake = Faker('he_IL')  # Hebrew locale

        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting FULL RANDOM anonymous booking test")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        # Random guests
        adults = random.randint(1, 3)
        children = random.randint(0, 2)
        infants = random.randint(0, 1)

        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f"🎯 Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results( hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()

        # Wait for form
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.ID, "checkout.personal_details_form.label_email"))
        )
        logging.info(" Order form visible — filling random data")

        # 🔀 Random personal details
        random_email = fake.email()
        random_phone = fake.phone_number().replace("-", "").replace("+", "")
        random_first_name = fake.first_name()
        random_last_name = fake.last_name()
        random_id = str(fake.random_int(min=100000000, max=999999999))
        random_note = fake.sentence(nb_words=6)

        self.order_page.set_email(random_email)
        self.order_page.set_phone(random_phone)
        self.order_page.set_first_name(random_first_name)
        self.order_page.set_last_name(random_last_name)
        self.order_page.set_id_number(random_id)

        self.order_page.click_terms_approval_checkbox_js()

        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys(random_note)

        # Click all special checkboxes (for fun)
        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        #  Random but valid-format card data
        card_number = "4580080111866879"
        expiry_month = str(random.randint(1, 12)).zfill(2)
        expiry_year = str(random.randint(2025, 2028))
        cvv = str(random.randint(100, 999))
        cardholder_name = fake.name()
        card_id_number = str(fake.random_int(min=1000000, max=9999999))

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number(card_number)
        self.order_page.select_expiry_month(expiry_month)
        self.order_page.select_expiry_year(expiry_year)
        self.order_page.set_cvv(cvv)
        self.order_page.set_cardholder_name(cardholder_name)
        self.order_page.set_id_number_card(card_id_number)

        self.order_page.switch_to_default_content()
        logging.info(" Full random booking flow finished (test mode)")
        # DO NOT Click the Order Button!!!
        # self.order_page.get_submit_button()

    def tearDown(self):
        if self.driver:
            try:
                result = getattr(self, "_outcome", None)  # PyCharm and unittest store results here
                if result is None:
                    logging.warning("No result object found. Skipping post_test_logging.")
                else:
                    self.post_test_logging(result)

                #  Log runtime duration
                try:
                    duration = self.driver.execute_script("""
                        const [start] = window.performance.getEntriesByName('selenium-start');
                        return Date.now() - start.startTime;
                    """)
                    logging.info(f" Test runtime: {int(duration)}ms")
                except Exception as e:
                    logging.warning(f" Could not get test runtime: {e}")

            except Exception as e:
                logging.warning(f"Logging failed during tearDown: {e}")

            logging.info("Closing browser (tearDown).")
            try:
                self.driver.quit()
            except Exception as e:
                logging.warning(f"Browser quit failed: {e}")

    if __name__ == "__main__":
        import unittest
        unittest.main()











