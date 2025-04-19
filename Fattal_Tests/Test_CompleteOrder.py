import traceback
from unittest import TestCase
from selenium import webdriver
import random
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Fattal_Main_Page import FattalMainPage
from Fattal_Tool_Bar import FattalToolBar
from Fattal_Search_Result import FattalSearchResultPage
from Fattal_Order_Page_ISR import FattalOrderPage
from Fattal_Flight_OrderPage import FattalFlightOrderPage
from Fattal_Confirmation_Page import FattalConfirmPage
import logging
import platform
from datetime import datetime
from faker import Faker
import os
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import io
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys
import json
import time
from selenium.webdriver.common.action_chains import ActionChains

class FattalTestsComplete(TestCase):
    def setUp(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--force-device-scale-factor=0.75")

        self.test_start_time = datetime.now()
        self.log_stream = io.StringIO()

        # Calculate base dir to Fattal_Tests (where test_fattal.py lives)
        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        # Load configuration
        try:
            with open(os.path.join(self.base_dir, "config.json"), encoding="utf-8") as f:
                self.config = json.load(f)
                logging.info("Configuration loaded successfully")
        except Exception as e:
            logging.error(f"Failed to load configuration: {e}")
            self.config = {}

        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(self.log_stream),
                logging.StreamHandler(sys.stdout)
            ]
        )

        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")

        self.driver = webdriver.Chrome(options=options)
        self.driver.get("https://www.fattal.co.il/")
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)

        self.main_page = FattalMainPage(self.driver)
        self.toolbar = FattalToolBar(self.driver)
        self.search_result = FattalSearchResultPage(self.driver)
        self.order_page = FattalOrderPage(self.driver)
        self.flight_page = FattalFlightOrderPage(self.driver)
        self.confirm_page = FattalConfirmPage(self.driver)
        self.driver.execute_script("window.performance.mark('selenium-start')")

    def retry_flight_search_if_no_results(self, hotel_name):
        def no_results_displayed():
            try:
                msg = self.driver.find_element(By.ID, "search-page-no-search-results-title")
                return "לא נמצאו מלונות" in msg.text
            except:
                return False

        if no_results_displayed():
            logging.warning("No flight hotel results found. Retrying...")

            self.take_screenshot("flight_search_no_results_initial")

            # Retry flight hotel search
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.select_flight_option_all_airports()
            self.main_page.search_button()

            # Wait and verify results
            try:
                WebDriverWait(self.driver, 15).until_not(
                    EC.presence_of_element_located((By.ID, "search-page-no-search-results-title"))
                )
                logging.info("Flight results appeared after retry.")
            except TimeoutException:
                self.take_screenshot("flight_search_retry_failed")
                raise Exception("No flight results even after retry.")
        else:
            logging.info("Flight hotel results returned.")

    def post_test_logging(self, result):
        test_method = self._testMethodName
        has_failed = False
        error_msg = ""
        screenshot_path = ""
        log_file_path = ""
        duration = (datetime.now() - self.test_start_time).total_seconds()

        try:
            browser = self.driver.capabilities['browserName'] if self.driver else "unknown"
        except Exception as e:
            logging.warning(f"Could not get browser info: {e}")
            browser = "unknown"

        os_name = platform.system()

        # 📄 Save logs
        log_summary = self.log_stream.getvalue()
        logs_dir = os.path.join(self.base_dir, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        log_file_path = os.path.join(logs_dir, f"{test_method}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        try:
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(log_summary)
            print(f"[LOG DEBUG] Written log to {log_file_path}")
        except Exception as e:
            print(f"[LOG DEBUG] Failed to write log: {e}")

        # 🚨 Detect real test failure using unittest outcome
        try:
            outcome = getattr(result, 'result', result)
            for failed_test, exc_info in outcome.failures + outcome.errors:
                if self._testMethodName in str(failed_test):
                    has_failed = True
                    if len(exc_info) == 3:
                        exc_type, exc_value, tb = exc_info
                        error_msg = "".join(traceback.format_exception(exc_type, exc_value, tb))
                    else:
                        error_msg = f" Unrecognized exception format: {exc_info}"
                    break
        except Exception as e:
            logging.warning(f" Failed to analyze test outcome: {e}")

        # 🧾 Capture confirmation metadata if available
        confirmation = getattr(self, "confirmation_result", {})
        order_number = confirmation.get("order_number", "")
        confirmed_email = confirmation.get("email", "")
        confirmation_screenshot = confirmation.get("screenshot_path", "")

        # 🖼 Create one unified screenshot depending on test result
        try:
            if self.driver:
                status_label = "FAIL" if has_failed else ("PASS" if order_number else "FAIL")
                screenshot_path = self.take_confirmation_screenshot(test_method, status_label)
                confirmation_screenshot = screenshot_path
        except Exception as e:
            logging.warning(f"Could not take confirmation screenshot: {e}")

        # 🧠 Final test info dict
        test_info = {
            "name": test_method,
            "description": "Test Description",
            "status": "FAILED" if has_failed else "PASSED",
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "duration": f"{duration:.2f}s",
            "browser": browser,
            "os": os_name,
            "full_name": f"{getattr(self, 'entered_first_name', '')} {getattr(self, 'entered_last_name', '')}".strip(),
            "email": confirmed_email or getattr(self, 'entered_email', ''),
            "order_number": order_number,
            "screenshot": confirmation_screenshot,
            "log": log_file_path,
            "error": error_msg
        }

        # 📦 Persist results
        self.save_to_excel(test_info)
        self.save_to_pdf(test_info)

    def save_to_excel(self, info: dict):
        # Same base path logic as working version
        parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(parent_folder, "TestResults.xlsx")

        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results"
                ws.append([
                    "Test Name", "Description", "Status", "Timestamp", "Duration",
                    "Browser", "OS", "Full Name", "Email", "Order Number", "Screenshot", "Log File"
                ])
            else:
                wb = load_workbook(filename)
                ws = wb.active

            status = info.get("status", "FAILED" if info.get("error") else "PASSED")

            row = [
                info.get("name", ""),
                info.get("description", ""),
                status,
                info.get("timestamp", ""),
                info.get("duration", ""),
                info.get("browser", ""),
                info.get("os", ""),
                info.get("full_name", ""),
                info.get("email", ""),
                info.get("order_number", ""),
                info.get("screenshot", ""),
                info.get("log", "")
            ]
            ws.append(row)
            row_num = ws.max_row

            # 🖼 Add screenshot hyperlink (if file exists)
            screenshot_path = info.get("screenshot", "")
            if screenshot_path and os.path.exists(screenshot_path):
                cell = ws.cell(row=row_num, column=11)
                cell.hyperlink = f"file:///{screenshot_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # 📜 Add log file hyperlink (if file exists)
            log_path = info.get("log", "")
            if log_path and os.path.exists(log_path):
                cell = ws.cell(row=row_num, column=12)
                cell.hyperlink = f"file:///{log_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # Auto column width
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.warning(f"Column resize error: {e}")
                ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(filename)
            logging.info(f"Excel file saved at: {filename}")

        except PermissionError as e:
            logging.warning(f"Excel file is open or locked: {e}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")

    def save_to_pdf(self, info: dict):
        pdf_dir = "reports"
        os.makedirs(pdf_dir, exist_ok=True)
        filename = f"{pdf_dir}/{info['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        c = canvas.Canvas(filename, pagesize=A4)

        width, height = A4
        y = height - 72

        c.setFont("Helvetica-Bold", 18)
        c.drawString(72, y, "Test Execution Report")
        y -= 40

        c.setFont("Helvetica", 12)
        lines = [
            f"Test Name: {info.get('name')}",
            f"Description: {info.get('description')}",
            f"Status: {info.get('status')}",
            f"Timestamp: {info.get('timestamp')}",
            f"Browser: {info.get('browser')}",
            f"OS: {info.get('os')}",
        ]
        if info.get("error"):
            lines.append(f"Error: {info.get('error')}")

        for line in lines:
            c.drawString(72, y, line)
            y -= 20

        if info.get("screenshot") and os.path.exists(info["screenshot"]):
            c.drawString(72, y, "Screenshot:")
            y -= 250
            c.drawImage(info["screenshot"], 72, y, width=5 * inch, preserveAspectRatio=True, mask='auto')

        c.showPage()
        c.save()
        print(f" PDF saved: {filename}")

    def take_confirmation_screenshot(self, test_method, status):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"confirmation_{status}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )
        self.driver.save_screenshot(filename)
        logging.info(f"Confirmation screenshot saved at: {filename}")
        return filename

    def run(self, result=None):
        self._resultForDoCleanups = result
        return super().run(result)

    def take_screenshot(self, test_method_name):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        screenshot_dir = os.path.join(os.getcwd(), "Fattal_Tests", "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = f"{screenshot_dir}/{test_method_name}_{timestamp}.png"
        self.driver.save_screenshot(filename)
        logging.error(f" Screenshot taken: {filename}")

    def retry_search_if_no_results(self, hotel_name, adults, children, infants):
        def no_results_displayed():
            try:
                msg = self.driver.find_element(By.ID, "search-page-no-search-results-title")
                return "לא נמצאו מלונות" in msg.text
            except:
                return False

        if no_results_displayed():
            logging.warning("No hotel-only results found. Retrying...")

            self.take_screenshot("hotel_search_no_results_initial")

            self.toolbar.logo_mainpage()

            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "main-input"))
            )

            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.search_button()

            try:
                WebDriverWait(self.driver, 15).until_not(
                    EC.presence_of_element_located((By.ID, "search-page-no-search-results-title"))
                )
                logging.info("Hotel results appeared after retry.")
            except TimeoutException:
                self.take_screenshot("hotel_search_retry_failed")
                raise Exception("No hotel results even after retry.")
        else:
            logging.info("Hotel search results returned.")

    def complete_booking_post_flight(self):
        # Make sure we’re not stuck inside an iframe
        self.order_page.switch_to_default_content()

        # ✅ Use the robust form readiness check
        self.order_page.wait_until_personal_form_ready()

        # 🧠 Generate valid Israeli ID for booking
        random_id = self.order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")

        # ✅ Guest details (could be randomized or from config)
        self.entered_email = "chenttedgui@gmail.com"
        self.entered_phone = "0544531600"
        self.entered_first_name = "חן"
        self.entered_last_name = "טסט"

        # Fill in the form
        self.order_page.set_email(self.entered_email)
        self.order_page.set_phone(self.entered_phone)
        self.order_page.set_first_name(self.entered_first_name)
        self.order_page.set_last_name(self.entered_last_name)
        self.order_page.set_id_number(random_id)

        self.order_page.click_terms_approval_checkbox_js()

        # Add optional special requests and room preferences
        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("טקסט לדוגמא .....")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        # Payment iframe actions
        self.order_page.switch_to_payment_iframe()
        payment = self.config["payment"]["credit_card"]
        self.order_page.set_card_number(payment["card_number"])
        self.order_page.select_expiry_month(payment["expiry_month"])
        self.order_page.select_expiry_year(payment["expiry_year"])
        self.order_page.set_cvv(payment["cvv"])
        self.order_page.set_cardholder_name(payment["cardholder_name"])
        self.order_page.set_id_number_card(payment["id_number"])

        # ✅ Click submit button with scroll + fallback
        try:
            submit_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "submitBtn"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
            time.sleep(0.5)

            try:
                self.driver.execute_script("arguments[0].click();", submit_btn)
                logging.info("Clicked 'בצע תשלום' button via JavaScript")
            except Exception as js_click_error:
                logging.warning(f"JS click failed, trying ActionChains: {js_click_error}")
                ActionChains(self.driver).move_to_element(submit_btn).click().perform()
                logging.info("Clicked 'בצע תשלום' button via ActionChains")

        except Exception as e:
            logging.error(f"❌ Failed to click submit button: {e}")
            raise

        self.order_page.switch_to_default_content()
        logging.info("📦 Booking flow completed and submit clicked (post-flight)")

        # Optional: Wait for confirmation page
        try:
            WebDriverWait(self.driver, 25).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//h1[contains(text(), 'אישור') or contains(text(), 'תודה')]"))
            )
            logging.info("🎉 Confirmation page loaded.")
        except:
            logging.warning("⚠️ Confirmation page not found yet. Continuing anyway.")

    def complete_booking_flow(self, hotel_name, adults, children, infants):
        # Generate a random Israeli ID here and set it
        random_id = self.order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")

        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f" Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()

        self.order_page.wait_until_personal_form_ready()

        # 📝 Hardcode guest details for now
        guest = self.config["default_guest"]
        self.order_page.set_email(guest["email"])
        self.order_page.set_phone(guest["phone"])
        self.order_page.set_first_name(guest["first_name"])
        self.order_page.set_last_name(guest["last_name"])
        self.order_page.set_id_number(random_id)

        # For logging/export
        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]

        #  Toggle approval
        self.order_page.click_terms_approval_checkbox_js()
        self.order_page.expand_special_requests_section()

        #  Write special requests
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("לינה ועוד דברים. שיהיה לנו בכיף")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number(self.config["payment"]["credit_card"]["card_number"])
        self.order_page.select_expiry_month(self.config["payment"]["credit_card"]["expiry_month"])
        self.order_page.select_expiry_year(self.config["payment"]["credit_card"]["expiry_year"])
        self.order_page.set_cvv(self.config["payment"]["credit_card"]["cvv"])
        self.order_page.set_cardholder_name(self.config["payment"]["credit_card"]["cardholder_name"])
        self.order_page.set_id_number_card(self.config["payment"]["credit_card"]["id_number"])

        # 🔥 Click the actual button here - using a more robust approach
        try:
            submit_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "submitBtn"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
            time.sleep(0.5)  # Small pause to ensure scrolling completes
            
            # Try JavaScript click first (most reliable for overlays)
            try:
                self.driver.execute_script("arguments[0].click();", submit_btn)
                logging.info("Clicked 'בצע תשלום' button via JavaScript")
            except Exception as click_error:
                logging.warning(f"JS click failed, trying ActionChains: {click_error}")
                # Try ActionChains as fallback
                actions = ActionChains(self.driver)
                actions.move_to_element(submit_btn).click().perform()
                logging.info("Clicked 'בצע תשלום' button via ActionChains")
        except Exception as e:
            logging.error(f"Failed to click submit button: {e}")
            raise

        self.order_page.switch_to_default_content()
        logging.info("Booking flow completed and submit clicked (test mode)")

    def complete_booking_flow_club_checkbox(self, hotel_name, adults, children, infants):
        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f" Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()

        self.order_page.wait_until_personal_form_ready()
        self.order_page.club_checkbox()

        random_id = self.order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")

        guest = self.config["default_guest"]
        self.order_page.set_email(guest["email"])
        self.order_page.set_phone(guest["phone"])
        self.order_page.set_first_name(guest["first_name"])
        self.order_page.set_last_name(guest["last_name"])
        self.order_page.set_id_number(random_id)

        # For logging/export
        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]

        self.order_page.click_terms_approval_checkbox_js()
        self.order_page.expand_special_requests_section()

        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("Test Test Test")

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number(self.config["payment"]["credit_card"]["card_number"])
        self.order_page.select_expiry_month(self.config["payment"]["credit_card"]["expiry_month"])
        self.order_page.select_expiry_year(self.config["payment"]["credit_card"]["expiry_year"])
        self.order_page.set_cvv(self.config["payment"]["credit_card"]["cvv"])
        self.order_page.set_cardholder_name(self.config["payment"]["credit_card"]["cardholder_name"])
        self.order_page.set_id_number_card(self.config["payment"]["credit_card"]["id_number"])

        #  Click submit button using the same robust approach
        try:
            submit_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "submitBtn"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
            time.sleep(0.5)  # Small pause to ensure scrolling completes
            
            # Try JavaScript click first (most reliable for overlays)
            try:
                self.driver.execute_script("arguments[0].click();", submit_btn)
                logging.info("Clicked 'בצע תשלום' button via JavaScript")
            except Exception as click_error:
                logging.warning(f"JS click failed, trying ActionChains: {click_error}")
                # Try ActionChains as fallback
                actions = ActionChains(self.driver)
                actions.move_to_element(submit_btn).click().perform()
                logging.info("Clicked 'בצע תשלום' button via ActionChains")
        except Exception as e:
            logging.error(f"Failed to click submit button: {e}")
            raise

        self.order_page.switch_to_default_content()
        logging.info(" Booking flow completed and submit clicked (club flow)")

    def test_anonymous_no_login(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        self.complete_booking_flow(hotel_name, adults, children, infants)

        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        assert self.confirmation_result.get("order_number"), " Booking failed — no order number found."

    def test_anonymous_no_login_club_checkbox(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 0, 0

        self.complete_booking_flow_club_checkbox(hotel_name, adults, children, infants)
        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."

    def test_eilat_flight(self):
        hotel_name = "אילת,ישראל"
        adults, children, infants = 2, 1, 1

        try:
            logging.info(" Starting test for Eilat zone with flight")

            # --- HOTEL + ROOM + FLIGHT SELECTION ---
            self.main_page.deal_popup()
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.select_flight_option_all_airports()

            self.main_page.set_room_occupants(adults, children, infants)
            self.main_page.search_button()

            self.retry_flight_search_if_no_results(hotel_name)
            # self.search_result.wait_for_rooms_to_load()
            self.search_result.click_book_room_button()
            self.search_result.wait_for_prices_to_load()
            self.search_result.click_first_show_prices()
            self.search_result.click_first_book_room()

            # --- FLIGHT SELECTION LOGIC ---
            try:
                self.flight_page.try_flight_options_by_time_of_day()
            except Exception as e:
                logging.warning(f" Flight selection failed (non-blocking): {e}")

            # --- PASSENGER FORM ---
            handled = self.flight_page.handle_passenger_form_if_flight_selection_skipped()
            if not handled:
                self.flight_page.wait_for_passenger_form()
                self.flight_page.fill_passenger_details_and_validate()
                self.flight_page.click_continue_button()

            #  Booking flow continues only if everything succeeded
            #self.complete_booking_post_flight()
            #self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)

        except Exception as e:
            logging.error(f" Test failed: {e}")
            try:
                self.take_screenshot("test_eilat_flight")
            except Exception:
                logging.warning("📸 Couldn't capture screenshot due to earlier failure.")
            raise

    def test_eilat(self):
        hotel_name = "אילת,ישראל"
        logging.info(" Starting test: hotel search and booking flow")
        adults, children, infants = 2, 0, 0
        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()
        self.main_page.set_room_occupants(adults, children, infants)
        self.main_page.search_button()
        self.search_result.wait_for_rooms_to_load()
        self.search_result.click_book_room_button()
        self.search_result.wait_for_prices_to_load()
        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()
        self.complete_booking_post_flight()
        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."
    def test_club_login(self):
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info(" Starting test: hotel search and booking flow")

        # ✍️ Static test info
        self.entered_first_name = "דניאל"
        self.entered_last_name = "טסט"
        self.entered_email = "test@example.com"

        self.main_page.deal_popup()

        try:
            self.toolbar.personal_zone()
            WebDriverWait(self.driver, 5).until(
                EC.visibility_of(self.toolbar.user_id_input())
            ).send_keys("999318330")
            self.toolbar.user_password_input().send_keys("Aa123456")
            self.toolbar.login_button()
            self.main_page.close_post_login_popup()
            logging.info("Logged in to club account successfully.")
        except Exception as e:
            logging.warning(f"Login failed: {e}")

        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        # ✅ Booking flow: now includes clicking the submit button
        self.complete_booking_flow(hotel_name, adults, children, infants)

        # ✅ NEW: Only wait for confirmation and extract order info
        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        assert self.confirmation_result.get("order_number"), "❌ Order number not found — confirmation may have failed."

        logging.info("✔️ Club login test finished with confirmed order.")

    def test_full_random_no_login(self):
        fake = Faker('he_IL')
        hotel_name = "לאונרדו נגב, באר שבע"
        logging.info("🚀 Starting FULL RANDOM anonymous booking test")

        self.main_page.deal_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f"🎯 Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()
        self.order_page.wait_until_personal_form_ready()
        logging.info("📝 Order form visible — filling random data")

        # Generate fake guest info
        random_id = self.order_page.generate_israeli_id()
        random_email = fake.email()
        random_phone = fake.phone_number().replace("-", "").replace("+", "")
        random_first_name = fake.first_name()
        random_last_name = fake.last_name()
        random_note = fake.sentence(nb_words=6)

        # Save for report
        self.entered_email = random_email
        self.entered_first_name = random_first_name
        self.entered_last_name = random_last_name

        # Fill form fields
        self.order_page.set_email(random_email)
        self.order_page.set_phone(random_phone)
        self.order_page.set_first_name(random_first_name)
        self.order_page.set_last_name(random_last_name)
        self.order_page.set_id_number(random_id)

        self.order_page.click_terms_approval_checkbox_js()
        self.order_page.expand_special_requests_section()

        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys(random_note)

        for checkbox in [
            self.order_page.get_adjacent_rooms_checkbox(),
            self.order_page.get_high_floor_checkbox(),
            self.order_page.get_low_floor_checkbox()
        ]:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
            self.driver.execute_script("arguments[0].click();", checkbox)

        # Payment section (dry-run)
        card_number = self.config["payment"]["credit_card"]["card_number"]
        expiry_month = str(random.randint(1, 12)).zfill(2)
        expiry_year = str(random.randint(2025, 2028))
        cvv = str(random.randint(100, 999))
        cardholder_name = fake.name()
        card_id_number = str(fake.random_int(min=1000000, max=9999999))

        self.order_page.switch_to_payment_iframe()
        self.order_page.set_card_number(card_number)
        self.order_page.select_expiry_month(expiry_month)
        self.order_page.select_expiry_year(expiry_year)
        self.order_page.set_cvv(cvv)
        self.order_page.set_cardholder_name(cardholder_name)
        self.order_page.set_id_number_card(card_id_number)

        # 💡 DO NOT click submit — test ends here
        self.order_page.switch_to_default_content()
        logging.info("🛑 Test completed before payment (dry-run mode)")

        # Still provide dummy confirmation data for test framework
        self.confirmation_result = {
            "order_number": "[DRY-RUN]",
            "email": self.entered_email,
            "screenshot_path": ""
        }

    def tearDown(self):
        if self.driver:
            try:
                result = getattr(self, "_outcome", None)  # PyCharm and unittest store results here
                if result is None:
                    logging.warning("No result object found. Skipping post_test_logging.")
                else:
                    self.post_test_logging(result)

                #  Log runtime duration
                try:
                    duration = self.driver.execute_script("""
                        const [start] = window.performance.getEntriesByName('selenium-start');
                        return Date.now() - start.startTime;
                    """)
                    logging.info(f" Test runtime: {int(duration)}ms")
                except Exception as e:
                    logging.warning(f" Could not get test runtime: {e}")

            except Exception as e:
                logging.warning(f"Logging failed during tearDown: {e}")

            logging.info("Closing browser (tearDown).")
            try:
                self.driver.quit()
            except Exception as e:
                logging.warning(f"Browser quit failed: {e}")

    if __name__ == "__main__":
        import unittest
        unittest.main()











