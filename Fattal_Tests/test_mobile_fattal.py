import io
import logging
import sys
import time
import traceback
from openpyxl.styles import  PatternFill
from time import sleep
import unittest
from selenium import webdriver
import functools

from selenium.common import TimeoutException

from Mobile_Fattal_Pages.Mobile_Fattal_Flight_Page import FattalFlightPageMobile
from Mobile_Fattal_Pages.Mobile_Toolbar_Fattal import FattalMobileToolBar
from Mobile_Fattal_Pages.Mobile_Fattal_Main_Page import FattalMainPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Search_Result_Page import FattalSearchResultPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Order_Page import FattalOrderPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_ConfirmPage import FattalMobileConfirmPage
from Mobile_Fattal_Pages.Mobile_Fattal_Deals_Packages import FattalDealsPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Customer_Contact_Page import FattalMobileCustomerSupport
from Mobile_Fattal_Pages.Mobile_Fattal_Join_Club_Page import FattalMobileClubJoinPage
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
HOTEL_NAME_TO_ID = {
    "לאונרדו נגב, באר שבע": "10048",
    "לאונרדו פלאזה אילת": "1051"
}


def retry_on_no_results(max_attempts=2):
    """
    Decorator to retry a test if a 'no results' situation is detected.
    """

    def decorator(func):
        @functools.wraps(func)
        def wrapper(self, *args, **kwargs):
            attempt = 0
            last_exception = None
            while attempt < max_attempts:
                try:
                    result = func(self, *args, **kwargs)
                    # If we reach here, check for the 'no results' UI (customize for your app)
                    if hasattr(self, "is_no_results_displayed") and self.is_no_results_displayed():
                        logging.warning(f"No results detected after attempt {attempt + 1}/{max_attempts}")
                        attempt += 1
                        continue
                    return result
                except Exception as e:
                    logging.warning(f"Attempt {attempt + 1} failed: {e}")
                    last_exception = e
                    attempt += 1
            # If we exhausted all retries
            if last_exception:
                raise last_exception
            else:
                raise Exception("Test failed after all retry attempts (no results).")

        return wrapper

    return decorator
class FattalMobileTests(unittest.TestCase):
    def save_order_for_cancellation(self, order_number: str):
        import json

        try:
            if not order_number:
                logging.warning("Order number is empty, not saving for cancellation.")
                return

            hotel_id = HOTEL_NAME_TO_ID.get(self.default_hotel_name, "UNKNOWN")

            order_entry = {
                "masterID": str(order_number),
                "hotelID": hotel_id
            }

            # Ensure path is relative to current test script
            current_dir = os.path.dirname(os.path.abspath(__file__))
            cancel_file = os.path.join(current_dir, "orders_to_cancel.json")

            if os.path.exists(cancel_file):
                try:
                    with open(cancel_file, "r", encoding="utf-8") as f:
                        existing_orders = json.load(f)
                        if not isinstance(existing_orders, list):
                            logging.warning("⚠️ JSON file does not contain a list. Resetting it.")
                            existing_orders = []
                except Exception as e:
                    logging.warning(f"⚠️ JSON file invalid — starting fresh. ({e})")
                    existing_orders = []
            else:
                existing_orders = []

            existing_orders.append(order_entry)

            with open(cancel_file, "w", encoding="utf-8") as f:
                json.dump(existing_orders, f, indent=2, ensure_ascii=False)

            logging.info(f"✅ Saved order {order_number} for cancellation (JSON included).")

        except Exception as e:
            logging.warning(f"⚠️ Could not save order for cancellation: {e}")

    def setUp(self):
        load_dotenv()
        self.log_stream = io.StringIO()
        # Reset logging configuration
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(self.log_stream),
                logging.StreamHandler(sys.stdout)
            ]
        )

        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        # Use a valid built-in Chrome emulation device (Pixel 2)
        mobile_emulation = {
            "deviceMetrics": {"width": 411, "height": 800, "pixelRatio": 3.0},
            "userAgent": "Mozilla/5.0 (Linux; Android 10; Pixel 2 XL) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36"
        }

        options = webdriver.ChromeOptions()
        options.add_experimental_option("mobileEmulation", mobile_emulation)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        # options.add_argument("--disable-infobars")
        # options.add_argument("--headless")
        # options.add_argument("--no-sandbox")
        # options.add_argument("--disable-dev-shm-usage")

        self.driver = webdriver.Chrome(options=options)

        # Override window size manually to make it taller 📸
        # Pixel 2 is normally 411x731, so let's go 411x1200
        self.driver.set_window_rect(x=620, y=0, width=411, height=950)

        self.test_start_time = datetime.now()
        self.driver.implicitly_wait(10)

        # Optional — force touch emulation for some mobile interactions
        self.driver.execute_cdp_cmd("Emulation.setTouchEmulationEnabled", {
            "enabled": True,
            "configuration": "mobile"
        })
        active_key = os.getenv("ENV_ACTIVE")
        self.driver.get(active_key)
        logging.info(f"Opened environment URL: {active_key}")
        # Load defaults from .env
        self.default_guest = {
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
            "first_name": os.getenv("DEFAULT_FIRST_NAME"),
            "last_name": os.getenv("DEFAULT_LAST_NAME")
        }
        self.default_guest_europe = {
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
            "first_name": os.getenv("DEFAULT_FIRST_NAME_ENGLISH"),
            "last_name": os.getenv("DEFAULT_LAST_NAME_ENGLISH")
        }
        self.default_hotel_name = os.getenv("DEFAULT_HOTEL_NAME")
        self.default_hotel_name_europe = os.getenv("DEFAULT_HOTEL_NAME_EUROPE")
        self.payment_card = {
            "card_number": os.getenv("PAYMENT_CARD_NUMBER"),
            "cardholder_name": os.getenv("PAYMENT_CARDHOLDER_NAME"),
            "expiry_month": os.getenv("PAYMENT_EXPIRY_MONTH"),
            "expiry_year": os.getenv("PAYMENT_EXPIRY_YEAR"),
            "cvv": os.getenv("PAYMENT_CVV"),
            "id_number": os.getenv("PAYMENT_ID_NUMBER")
        }

        # Page object initialization
        self.mobile_main_page = FattalMainPageMobile(self.driver)
        self.mobile_search_page = FattalSearchResultPageMobile(self.driver)
        self.mobile_order_page = FattalOrderPageMobile(self.driver)
        self.mobile_toolbar = FattalMobileToolBar(self.driver)
        self.mobile_confirm = FattalMobileConfirmPage(self.driver)
        self.mobile_deals_page = FattalDealsPageMobile(self.driver)
        self.mobile_flight_page = FattalFlightPageMobile(self.driver)
        self.mobile_customer_support = FattalMobileCustomerSupport(self.driver)
        self.mobile_club_join_page = FattalMobileClubJoinPage(self.driver)

    def soft_assert(self, condition, msg, errors_list):
        """
        A soft assert will check the condition. If it fails, it will log the error and add it to a list.
        The test will continue to run without failing the test immediately.
        """
        if not condition:
            logging.error(f"Soft Assertion Failed: {msg}")  # Log the failure
            errors_list.append(msg)  # Add the failure message to the list of errors

    def post_test_logging_mobile(self, result):
        test_method = self._testMethodName
        has_failed = False
        error_msg = ""
        log_file = ""
        duration = (datetime.now() - self.test_start_time).total_seconds()

        # ── Raw Logs ───────────────────────────────────────
        logs_dir = os.path.join(self.base_dir, "logs_mobile")
        os.makedirs(logs_dir, exist_ok=True)
        log_file = os.path.join(logs_dir, f"{test_method}_{datetime.now():%Y-%m-%d_%H-%M-%S}.log")
        try:
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(self.log_stream.getvalue())
        except Exception as e:
            logging.warning(f"[MOBILE] Could not write log file: {e}")

        # ── Result Analysis ────────────────────────────────
        try:
            outcome = getattr(result, 'result', result)
            for failed_test, exc_info in outcome.failures + outcome.errors:
                if self._testMethodName in str(failed_test):
                    has_failed = True
                    if len(exc_info) == 3:
                        exc_type, exc_value, tb = exc_info
                        error_msg = "".join(traceback.format_exception(exc_type, exc_value, tb))
                    else:
                        error_msg = str(exc_info)
                    break
        except Exception as e:
            logging.warning(f"[MOBILE] Failed to parse result: {e}")

        # ── Confirmation Data ──────────────────────────────
        conf = getattr(self, "confirmation_result", {}) or {}
        order_no = conf.get("order_number", "")
        email = conf.get("email", "") or getattr(self, "entered_email", "")

        # ── Screenshots ───────────────────────────────────
        # This picks up your screenshot paths if they were set
        confirmation_screenshot = getattr(self, "screenshot_confirmation", "")
        error_screenshot = getattr(self, "screenshot_error", "")

        # ── Final Struct & Save ────────────────────────────
        status = "FAILED" if has_failed else "PASSED"
        info = {
            "name": test_method,
            "description": getattr(self, "test_description", ""),
            "status": status,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "duration": f"{duration:.2f}s",
            "browser": "unknown",  # Update if needed
            "os": "unknown",  # Update if needed
            "full_name": f"{getattr(self, 'entered_first_name', '')} {getattr(self, 'entered_last_name', '')}".strip(),
            "email": email,
            "order_number": order_no,
            "id_number": getattr(self, "entered_id_number", ""),
            "confirmation_screenshot": confirmation_screenshot,
            "error_screenshot": error_screenshot,
            "log": log_file,
            "error": error_msg if has_failed else ""
        }

        # Log soft assertion errors if they exist
        if self.soft_assert_errors:
            logging.info(f"Soft Assertion Errors found: {len(self.soft_assert_errors)}")
            for error in self.soft_assert_errors:
                logging.info(f"Soft Assertion Failure: {error}")

        # Optional: Save to Excel/HTML if needed
        self.save_to_excel(info)
        self.save_to_html(info)

    def save_to_excel(self, info: dict):
        SCREENSHOT_LABEL = "📷 Screenshot"
        LOG_LABEL = "🧾 Log File"
        parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(parent_folder, "TestResults.xlsx")

        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results"
                ws.append([
                    "Test Name", "Order Number", "ID Number", "Description", "Status", "Timestamp", "Duration",
                    "Browser", "OS", "Full Name", "Email",
                    "Room Screenshot", "Payment Screenshot", "Confirmation Screenshot", "Log File"
                ])
            else:
                wb = load_workbook(filename)
                ws = wb.active

            status = "FAILED" if info.get("error") else "PASSED"

            row = [
                info.get("name", ""),
                info.get("order_number", ""),
                info.get("id_number", ""),
                info.get("description", ""),
                status,
                info.get("timestamp", ""),
                info.get("duration", ""),
                info.get("browser", ""),
                info.get("os", ""),
                info.get("full_name", ""),
                info.get("email", ""),
                getattr(self, "screenshot_room_selection", ""),
                getattr(self, "screenshot_payment_stage", ""),
                info.get("error_screenshot" if info.get("status") == "FAILED" else "confirmation_screenshot", ""),
                info.get("log", "")
            ]
            ws.append(row)
            row_num = ws.max_row

            # 🎨 Color row based on result
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row_fill = red_fill if status == "FAILED" else green_fill

            for col_idx in range(1, 16):  # Adjusted to match 15 columns
                ws.cell(row=row_num, column=col_idx).fill = row_fill

            # 📎 Add hyperlinks for screenshots (Room, Payment, Confirmation)
            for col_idx in [12, 13, 14]:  # Screenshot columns
                screenshot_path = row[col_idx - 1]
                if screenshot_path and os.path.exists(screenshot_path):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = SCREENSHOT_LABEL
                    cell.hyperlink = f"file:///{screenshot_path.replace(os.sep, '/')}"
                    cell.font = Font(color="0000EE", underline="single")

            # 📎 Add hyperlink for log file
            log_path = info.get("log", "")
            if log_path and os.path.exists(log_path):
                cell = ws.cell(row=row_num, column=15)
                cell.value = LOG_LABEL
                cell.hyperlink = f"file:///{log_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # ↔️ Auto-fit column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.warning(f"Column resize error: {e}")
                ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(filename)
            logging.info(f"Excel file saved at: {filename}")

        except PermissionError as e:
            logging.warning(f"Excel file is open or locked: {e}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")

    def save_to_html(self, info: dict):
        import time
        from datetime import datetime
        import os

        html_dir = os.path.join(self.base_dir, "html_reports")
        os.makedirs(html_dir, exist_ok=True)

        dashboard_path = os.path.join(html_dir, "TestDashboard.html")

        def image_tag(path):
            if path and os.path.exists(path):
                rel_path = os.path.relpath(path, start=html_dir).replace(os.sep, '/')
                return f'<img src="{rel_path}" onclick="openModal(this.src)" style="max-height:120px; border:1px solid #ccc; margin:5px; cursor: pointer;" />'
            return "<em>No screenshot</em>"

        room_img = image_tag(getattr(self, 'screenshot_room_selection', ''))
        pay_img = image_tag(getattr(self, 'screenshot_payment_stage', ''))
        confirm_img = image_tag(
            info.get('error_screenshot') if info.get('status') == 'FAILED' else info.get('confirmation_screenshot'))

        test_type_class = info.get("test_type", "desktop").lower()

        try:
            ts_ms = int(datetime.strptime(info.get('timestamp'), '%Y-%m-%d %H:%M:%S').timestamp() * 1000)
        except:
            ts_ms = int(time.time() * 1000)

        log_file_rel = ""
        if info.get("log") and os.path.exists(info["log"]):
            log_file_rel = os.path.relpath(info["log"], start=html_dir).replace(os.sep, '/')

        html_entry = f"""
        <div class=\"test-block {test_type_class}\" data-timestamp=\"{ts_ms}\">
            <h2>{info.get('name')} — <span class=\"{'fail' if info.get('status') == 'FAILED' else 'pass'}\">{info.get('status')}</span></h2>
            <p><strong>Description:</strong> {info.get('description')}</p>
            <p><strong>Timestamp:</strong> {info.get('timestamp')} | <strong>Duration:</strong> {info.get('duration')}</p>
            <p><strong>Browser:</strong> {info.get('browser')} | <strong>OS:</strong> {info.get('os')}</p>
            <p><strong>Guest:</strong> {info.get('full_name')} | <strong>Email:</strong> {info.get('email')}</p>
            <p><strong>Order #:</strong> {info.get('order_number')} | <strong>ID:</strong> {info.get('id_number')}</p>
            <p><strong>Log File:</strong> <a href=\"{log_file_rel}\" target=\"_blank\">View Log</a></p>
            {'<p style="color:red;"><strong>Error:</strong> ' + info['error'] + '</p>' if info.get('error') else ''}
            <div class=\"grid\">
                <div><h4>Room</h4>{room_img}</div>
                <div><h4>Payment</h4>{pay_img}</div>
                <div><h4>{'Failure Screenshot' if info.get('status') == 'FAILED' else 'Confirmation'}</h4>{confirm_img}</div>
            </div>
            <hr>
        </div>
        """

        if not os.path.exists(dashboard_path):
            html_start = """<!DOCTYPE html>
    <html lang=\"en\">
    <head>
        <meta charset=\"UTF-8\">
        <title>Fattal Selenium Test Dashboard</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            .pass { color: green; font-weight: bold; }
            .fail { color: red; font-weight: bold; }
            .grid { display: flex; gap: 20px; margin-top: 10px; flex-wrap: wrap; }
            .grid div { flex: 1; min-width: 250px; }
            .test-block { margin-bottom: 40px; }
            hr { border: 1px dashed #ccc; }
            img { display: block; max-width: 100%; }
            img:hover {
                transform: scale(1.05);
                transition: transform 0.2s ease-in-out;
                box-shadow: 0 0 10px rgba(0,0,0,0.4);
            }
            .summary-bar {
                background-color: #f5f5f5;
                border: 1px solid #ccc;
                padding: 10px;
                margin-bottom: 20px;
            }
            .filters { margin-bottom: 20px; }
            .filters label { margin-right: 15px; cursor: pointer; font-weight: normal; }
            .filters select, .filters button { margin-right: 10px; }
            .filters input[type=\"checkbox\"] {
                margin-right: 5px;
                transform: scale(1.1);
                vertical-align: middle;
            }
            .modal {
                display: none;
                position: fixed;
                z-index: 999;
                padding-top: 40px;
                left: 0; top: 0;
                width: 100%; height: 100%;
                overflow: auto;
                background-color: rgba(0,0,0,0.85);
            }
            .modal-content {
                margin: auto;
                display: block;
                max-width: 90vw;
                max-height: 80vh;
                object-fit: contain;
                border-radius: 8px;
            }
            .modal-content, .close {
                animation-name: zoom;
                animation-duration: 0.3s;
            }
            @keyframes zoom {
                from {transform: scale(0)} to {transform: scale(1)}
            }
            .close {
                position: absolute;
                top: 15px;
                right: 35px;
                color: #fff;
                font-size: 40px;
                font-weight: bold;
                cursor: pointer;
            }
            @media (max-width: 768px) {
                .modal-content {
                    max-width: 95%;
                    transform: none !important;
                }
                img:hover {
                    transform: none;
                }
            }
        </style>
        <script>
            function updateSummary() {
                let total = document.querySelectorAll('.test-block').length;
                let passed = document.querySelectorAll('.test-block .pass').length;
                let failed = document.querySelectorAll('.test-block .fail').length;
                document.getElementById('summary-total').textContent = total;
                document.getElementById('summary-passed').textContent = passed;
                document.getElementById('summary-failed').textContent = failed;
            }

            function applyFilters() {
                const showPassed = document.getElementById('filter-pass').checked;
                const showFailed = document.getElementById('filter-fail').checked;
                const showMobile = document.getElementById('filter-mobile').checked;
                const showDesktop = document.getElementById('filter-desktop').checked;
                const range = document.getElementById('date-range').value;
                const now = new Date();
                let minTime = 0;

                if (range === 'today') {
                    minTime = new Date(now.getFullYear(), now.getMonth(), now.getDate()).getTime();
                } else if (range === 'yesterday') {
                    minTime = new Date(now.getFullYear(), now.getMonth(), now.getDate() - 1).getTime();
                } else if (range === 'last7') {
                    minTime = now.getTime() - 7 * 24 * 60 * 60 * 1000;
                }

                document.querySelectorAll('.test-block').forEach(block => {
                    const isPass = block.querySelector('span').classList.contains('pass');
                    const isFail = block.querySelector('span').classList.contains('fail');
                    const isMobile = block.classList.contains('mobile');
                    const isDesktop = block.classList.contains('desktop');
                    const blockTime = parseInt(block.getAttribute('data-timestamp'), 10);

                    let visible = ((showPassed && isPass) || (showFailed && isFail)) &&
                                  ((showMobile && isMobile) || (showDesktop && isDesktop)) &&
                                  (range === 'all' || blockTime >= minTime);

                    block.style.display = visible ? 'block' : 'none';
                });
            }

            function openModal(imgSrc) {
                const modal = document.getElementById("screenshotModal");
                const modalImg = document.getElementById("modalImage");
                modal.style.display = "block";
                modalImg.src = imgSrc;
            }

            function closeModal() {
                document.getElementById("screenshotModal").style.display = "none";
            }

            window.onload = function() {
                updateSummary();
                applyFilters();
            };
        </script>
    </head>
    <body>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h1 style="margin: 0;">🧪 Fattal Selenium Test Dashboard</h1>
            <img src="fattal_logo.png" alt="Fattal Logo" style="height: 150px;" />
        </div>
        <div class="summary-bar">
            <strong>Total Tests:</strong> <span id="summary-total">-</span> |
            <strong style="color:green;">Passed:</strong> <span id="summary-passed">-</span> |
            <strong style="color:red;">Failed:</strong> <span id="summary-failed">-</span>
        </div>
        <div class="filters">
            <label><input type="checkbox" id="filter-pass" checked onchange="applyFilters()">✅ Passed</label>
            <label><input type="checkbox" id="filter-fail" checked onchange="applyFilters()">❌ Failed</label>
            <label><input type="checkbox" id="filter-desktop" checked onchange="applyFilters()">💻 Desktop</label>
            <label><input type="checkbox" id="filter-mobile" checked onchange="applyFilters()">📱 Mobile</label>
            <label><strong>Date:</strong></label>
            <select id="date-range" onchange="applyFilters()">
                <option value="all">All</option>
                <option value="today">Today</option>
                <option value="yesterday">Yesterday</option>
                <option value="last7">Last 7 Days</option>
            </select>
        </div>
    """
            with open(dashboard_path, "w", encoding="utf-8") as f:
                f.write(html_start)

        with open(dashboard_path, "a", encoding="utf-8") as f:
            f.write(html_entry)

        with open(dashboard_path, "a", encoding="utf-8") as f:
            f.write("""
            <div id=\"screenshotModal\" class=\"modal\" onclick=\"closeModal()\">
              <span class=\"close\">&times;</span>
              <img class=\"modal-content\" id=\"modalImage\">
            </div>
        </body>
        </html>
        """)

    def is_no_results_displayed(self):
        try:
            # Customize selector according to your "No results" element!
            return self.driver.find_element(By.XPATH, "//div[contains(text(), 'לא נמצאו תוצאות')]").is_displayed()
        except Exception:
            return False

    def fill_guest_details(self, guest=None, email=None, phone=None, first_name=None, last_name=None):
        """
        Fills guest details into the order page.
        You can either pass a full `guest` dictionary (with keys: email, phone, first_name, last_name)
        OR use individual arguments to override defaults.
        """
        if guest:
            email = guest.get("email", email)
            phone = guest.get("phone", phone)
            first_name = guest.get("first_name", first_name)
            last_name = guest.get("last_name", last_name)

        # Defaults fallback
        email = email or "chenttedgui@gmail.com"
        phone = phone or "0544531600"
        first_name = first_name or "חן"
        last_name = last_name or "טסט"

        try:
            # Find email field and scroll into view
            email_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "checkout-form-field-input_email"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", email_input)
            time.sleep(0.5)  # Short pause after scroll

            # Wait until clickable
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "checkout-form-field-input_email"))
            )
            logging.info("✅ Email field is ready for input.")
        except TimeoutException:
            self.take_screenshot("email_not_clickable_timeout")
            raise Exception("❌ Email input not clickable after timeout.")

        # Fill the fields
        self.mobile_order_page.set_email(email)
        self.mobile_order_page.set_phone(phone)
        self.mobile_order_page.set_first_name(first_name)
        self.mobile_order_page.set_last_name(last_name)

        # For logging/export
        self.entered_email = email
        self.entered_first_name = first_name
        self.entered_last_name = last_name

    def fill_payment_details_from_config(self):
        """
        Fills payment details using the credit card information from config.json
        """
        try:
            cardholder_name = self.payment_card["cardholder_name"]
            card_number = self.payment_card["card_number"]
            expiry_month = self.payment_card["expiry_month"]
            expiry_year = self.payment_card["expiry_year"]
            cvv = self.payment_card["cvv"]
            id_number = self.payment_card["id_number"]
            
            logging.info("Using credit card details from config.json")
            
            # Find and switch to the iframe
            iframe = self.driver.find_element(By.ID, "paymentIframe")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", iframe)
            self.driver.switch_to.frame(iframe)
            # Fill Cardholder Name
            cardholder_input = self.driver.find_element(By.ID, "card_holder_name_input")
            cardholder_input.clear()
            cardholder_input.send_keys(cardholder_name)
            # Fill Card Number
            card_number_input = self.driver.find_element(By.ID, "credit_card_number_input")
            card_number_input.clear()
            card_number_input.send_keys(card_number)

            # Select Expiry Month
            month_select = Select(self.driver.find_element(By.ID, "date_month_input"))
            month_select.select_by_visible_text(expiry_month)
            
            # Select Expiry Year
            year_select = Select(self.driver.find_element(By.ID, "date_year_input"))
            year_select.select_by_visible_text(expiry_year)
            
            # Fill CVV
            cvv_input = self.driver.find_element(By.ID, "cvv_input")
            cvv_input.clear()
            cvv_input.send_keys(cvv)
            
            # Fill ID Number
            id_input = self.driver.find_element(By.ID, "id_number_input")
            id_input.clear()
            id_input.send_keys(id_number)
            
            logging.info("Credit card details from config applied successfully")
            
        except Exception as e:
            logging.error(f"Failed to apply credit card details from config: {e}")
            raise
        finally:
            # Switch back to default content
            self.driver.switch_to.default_content()

    def take_confirmation_screenshot(self, test_method, status):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"confirmation_{status}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )

        try:
            # Locate the element you want to anchor the screenshot from
            element = self.driver.find_element(By.ID, "thank-you-page-top-bar-text")

            # Get element's Y position and scroll there minus a bit (padding offset)
            y_position = element.location['y']
            self.driver.execute_script("window.scrollTo(0, arguments[0]);", max(y_position - 50, 0))

            time.sleep(1)  # Let rendering stabilize

            self.driver.save_screenshot(filename)
            logging.info(f" Screenshot anchored with manual scroll saved at: {filename}")
        except Exception as e:
            logging.warning(f" Could not scroll to confirmation element: {e}")
            try:
                self.driver.save_screenshot(filename)
                logging.info(f" Screenshot (fallback full view) saved at: {filename}")
            except Exception as inner_e:
                logging.error(f" Screenshot failed: {inner_e}")
                raise

        return filename

    def take_confirmation_screenshot_renew_membership(self, test_method, status):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"confirmation_{status}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )

        try:
            # Scroll to the container div for the confirmation message
            element = self.driver.find_element(By.CSS_SELECTOR, "div.sc-778b26e1-1.hFoaKR")

            # Scroll to the element's Y position with a slight offset
            y_position = element.location['y']
            self.driver.execute_script("window.scrollTo(0, arguments[0]);", max(y_position - 50, 0))

            time.sleep(1)  # Allow time for scroll and render

            self.driver.save_screenshot(filename)
            logging.info(f" Screenshot anchored to confirmation container saved at: {filename}")
        except Exception as e:
            logging.warning(f" Could not scroll to confirmation container: {e}")
            try:
                self.driver.save_screenshot(filename)
                logging.info(f" Screenshot (fallback full view) saved at: {filename}")
            except Exception as inner_e:
                logging.error(f" Screenshot failed: {inner_e}")
                raise

        return filename

    def run(self, result=None):
        self._test_result_for_teardown = result  # Save for later
        return super().run(result)

    def take_screenshot(self, test_method_name):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"{test_method_name}_{timestamp}.png"
        )
        self.driver.save_screenshot(filename)
        logging.error(f" Screenshot taken: {filename}")

    def take_stage_screenshot(self, label: str):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"{label}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )
        try:
            self.driver.save_screenshot(filename)
            logging.info(f"📸 Screenshot '{label}' saved at: {filename}")
            setattr(self, f"screenshot_{label}", filename)  # Dynamically set
        except Exception as e:
            logging.warning(f"❌ Could not take screenshot for '{label}': {e}")

    def test_mobile_join_fattal_and_friends_form(self):
        self.soft_assert_errors = []

        self.test_description = "הצטפרות למועדון דרך טופס"
        logging.info("Starting test: Join Fattal Club")

        try:
            # Step 1: Prepare data
            guest = self.default_guest
            first_name = guest["first_name"]
            last_name = guest["last_name"]
            email = guest["email"]
            phone = guest["phone"]
            birthdate = "01-01-1990"
            password = "Aa123456"
            id_number = self.mobile_order_page.generate_israeli_id()
            self.entered_id_number = id_number
            logging.info(f"Generated ID for club registration: {id_number}")
            self.mobile_main_page.close_war_popup()

            # Step 2: Navigate to club join screen
            self.mobile_toolbar.click_more_tab_mobile()
            self.mobile_toolbar.click_fattal_friends_club_tab()
            self.mobile_club_join_page.click_join_fattal_friends_button()

            # Step 3: Fill form
            self.mobile_club_join_page.fill_join_fattal_club_form(
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                id_number=id_number,
                birthdate=birthdate,
                password=password
            )

            # Step 4: Assert values
            self.mobile_club_join_page.assert_input_value("שם פרטי", first_name)
            self.mobile_club_join_page.assert_input_value("שם משפחה", last_name)
            self.mobile_club_join_page.assert_input_value("כתובת דוא״ל", email)
            self.mobile_club_join_page.assert_input_value("מספר טלפון נייד", phone)
            self.mobile_club_join_page.assert_input_value("תאריך לידה", birthdate)
            self.mobile_club_join_page.assert_input_value("מספר תעודת זהות", id_number)
            self.mobile_club_join_page.assert_input_value("בחרו סיסמא", password)

            logging.info("Fattal Club form filled and validated successfully.")
            self.mobile_club_join_page.click_accept_terms_checkbox()
            sleep(15)

            # Step 5: Payment
            self.fill_payment_details_from_config()
            self.mobile_order_page.click_payment_submit_button()
            self.mobile_toolbar.close_any_club_popup()

            sleep(3)
            # Step 7: Take confirmation screenshot
            self.confirmation_screenshot_path = self.take_confirmation_screenshot_renew_membership(self._testMethodName,
                                                                                                   "success")
            setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

        except Exception as e:
            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
            screenshot_dir = os.path.join(self.base_dir, "Screenshots")
            os.makedirs(screenshot_dir, exist_ok=True)
            screenshot_path = os.path.join(screenshot_dir, f"join_club_test_fail_{timestamp}.png")
            self.driver.save_screenshot(screenshot_path)
            logging.exception(f"Join Fattal Club test failed. Screenshot saved: {screenshot_path}")
            raise

        # Final check for soft assert errors
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

    # def test_mobile_contact_form(self):
    #     self.soft_assert_errors = []
    #
    #     self.test_description = "בדיקת טופס צור קשר"
    #     logging.info("Starting test: Customer Support Ticket")
    #
    #     try:
    #         guest = self.default_guest
    #         first_name = guest["first_name"]
    #         last_name = guest["last_name"]
    #         phone = guest["phone"]
    #         email = guest["email"]
    #         id_number = self.mobile_order_page.generate_israeli_id()
    #         self.entered_first_name = first_name
    #         self.entered_last_name = last_name
    #         self.entered_email = email
    #         self.entered_id_number = id_number
    #         message = "בדיקה אוטומטית של צור קשר דרך המובייל"
    #         hotel_name = "הרודס בוטיק אילת"
    #
    #         # Step 1: Open the contact form
    #         self.mobile_toolbar.click_more_tab_mobile()
    #         self.mobile_toolbar.click_contact_us_button_mobile()
    #         self.mobile_customer_support.click_send_us_inquiry_button()
    #
    #         # Step 2: Select 'נושא' dropdown
    #         self.mobile_customer_support.select_dropdown_by_label("נושא")
    #
    #         # Step 3: Fill form inputs
    #         self.mobile_customer_support.fill_basic_contact_fields(
    #             first_name, last_name, id_number, phone, email, message, accept_marketing=False
    #         )
    #
    #         # Step 4: Select hotel name
    #         self.mobile_customer_support.select_dropdown_by_label("שם המלון", option_text=hotel_name)
    #
    #         # 🖼️ Screenshot after form is filled
    #         self.take_stage_screenshot("contact_form_filled")
    #
    #         # Step 5: Verify form data
    #         self.mobile_customer_support.assert_form_data_matches_input(
    #             first_name=first_name,
    #             last_name=last_name,
    #             id_number=id_number,
    #             phone=phone,
    #             email=email,
    #             message=message
    #         )
    #
    #         # ✅ Soft assert success
    #         self.soft_assert(True, "Contact form submitted and verified successfully.", self.soft_assert_errors)
    #         logging.info("Test completed successfully.")
    #
    #     except Exception as e:
    #         timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
    #         screenshot_dir = os.path.join(self.base_dir, "Screenshots")
    #         os.makedirs(screenshot_dir, exist_ok=True)
    #         screenshot_path = os.path.join(
    #             screenshot_dir, f"test_failure_{timestamp}.png"
    #         )
    #         self.driver.save_screenshot(screenshot_path)
    #         logging.exception(f"Test failed. Screenshot saved: {screenshot_path}")
    #         raise
    #
    #     # ❗ Final check for soft assert errors
    #     if self.soft_assert_errors:
    #        logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))
    def test_mobile_booking_with_club_login(self):
        self.save_for_cancellation = False  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        hotel_name = self.default_hotel_name
        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_REGULAR_ID"),
            "password": os.getenv("CLUB_REGULAR_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")

        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5: Choose Room and click it
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        # Step 6: Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.set_first_name("Chen")
        self.mobile_order_page.set_last_name("Test")
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()

        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_anonymous_user(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי"
        hotel_name = self.default_hotel_name
        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()
        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        #Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        #Step 6 : Order Page
        #self.mobile_order_page.click_room_selection_summary()

        self.mobile_order_page.wait_until_personal_form_ready()

        #Order Details
        self.take_stage_screenshot("payment_stage")
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  #  Save for logging/export

        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_anonymous_join_fattal_and_friends(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי + הצטפרות למועדון"

        hotel_name = self.default_hotel_name

        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()
        # Step 6 : Order Page
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        # Order Details
        self.fill_guest_details(guest=self.default_guest)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_club_member_eilat_with_flight(self):
        self.save_for_cancellation = False
        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר עם מועדון פעיל + טיסות"
        hotel_name = "אילת, ישראל"
        logging.info("Starting mobile booking test for Eilat including flights...")
        user = {
            "id": os.getenv("CLUB_REGULAR_ID"),
            "password": os.getenv("CLUB_REGULAR_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")

        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        #self.mobile_main_page.select_date_range_two_months_ahead()
        # Step 2: Select exact date range instead of the dynamic one
        self.mobile_main_page.select_date_range_two_months_ahead_eilat(stay_length=5)

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        # Step 4: Enable flight option (🆕 use the mobile version of the method!)
        self.mobile_main_page.select_flight_option_all_airports()
        self.mobile_main_page.click_mobile_search_button()
        # 🔁 Handle "no results" and fallback if needed
        self.mobile_search_page.handle_no_search_results_and_choose_alternative()
        # Step 5: Handle results and fallback suggestions if needed
        self.mobile_search_page.handle_search_flow_with_fallback(self)

        # Step 7: Select flight or continue
        #self.mobile_flight_page.try_flight_options_by_time_of_day()
        sleep(5)
        # Step 8: Passenger form
        self.mobile_flight_page.fill_adult_passenger_details()
        self.mobile_flight_page.click_continue_button()

        # Step 9: Payment form
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    #@retry_on_no_results(max_attempts=3)
    def test_mobile_booking_anonymous_region_eilat(self):
        self.save_for_cancellation = False  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי דרך אזור מלונות אילת"
        hotel_name = "אילת, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead_eilat(stay_length=5)
        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        # Step 3: Search Vacation
        self.mobile_main_page.click_room_continue_button(),
        self.mobile_main_page.click_mobile_search_button()
        # 🔁 Handle "no results" and fallback if needed
        self.mobile_search_page.handle_no_search_results_and_choose_alternative()
        # Step 5: Handle results and fallback suggestions if needed
        self.mobile_search_page.handle_search_flow_with_fallback(self)
        # Step 6 : Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        # Order Details
        sleep(5)
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_fattal_gift3(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי + קופון 1 של פתאל גיפטס באילת"
        hotel_name = "תל אביב, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead(stay_length=5)
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.click_mobile_search_button()

        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(5)

        # Apply all 3 gift codes
        raw_gifts = [os.getenv("GIFT1"), os.getenv("GIFT2"), os.getenv("GIFT3")]
        gifts = [code.strip() for code in raw_gifts if code and code.strip().isdigit()]
        failed_gifts = []

        for code in gifts:
            logging.info(f"Applying gift coupon: '{code}'")
            self.mobile_order_page.apply_checkout_coupon(code)
            sleep(3)
            if not self.mobile_order_page.is_coupon_applied_successfully(code):
                logging.error(f"Gift code '{code}' was not applied successfully!")
                failed_gifts.append(code)
            else:
                logging.info(f"Gift code '{code}' applied successfully.")

        assert not failed_gifts, f"The following gift codes failed to apply: {', '.join(failed_gifts)}"

        self.fill_payment_details_from_config()
        self.mobile_order_page.click_payment_submit_button()

        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_fattal_gift1(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי +3 קופונים של פתאל גיפטס באילת"
        hotel_name = "תל אביב, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead(stay_length=5)
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.click_mobile_search_button()

        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)

        gift_code = os.getenv("GIFT4", "").strip()
        assert gift_code and gift_code.isdigit(), "GIFT4 is missing or invalid in environment config"

        logging.info(f"Applying single gift coupon: '{gift_code}'")
        self.mobile_order_page.apply_checkout_coupon(gift_code)
        sleep(15)

        applied_successfully = self.mobile_order_page.is_coupon_applied_successfully(gift_code)
        assert applied_successfully, f"Gift coupon '{gift_code}' failed to apply."

        self.fill_payment_details_from_config()
        self.mobile_order_page.click_payment_submit_button()

        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_club_member_club_renew_expired(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר מועדון בסטטוס פג תוקף"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_RENEW_ID"),
            "password": os.getenv("CLUB_RENEW_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")# For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"


        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_club_member_club_renew_about_to_expire(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר מועדון בסטטוס עומד לפוג"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_ABOUT_EXPIRE_ID"),
            "password": os.getenv("CLUB_ABOUT_EXPIRE_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()

            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()
        self.mobile_toolbar.handle_membership_renewal_popup()
        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_club_renew_expired_form(self):

        self.soft_assert_errors = []

        self.test_description = "חידוש מועדון דרך טופס"
        hotel_name = self.default_hotel_name
        logging.info("Starting test: Club renew expired form (mobile)")
        self.mobile_main_page.close_war_popup()

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_ABOUT_EXPIRE_ID_FORM"),
            "password": os.getenv("CLUB_ABOUT_EXPIRE_PASSWORD_FORM")
        }

        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")

        # 📊 For reporting
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        self.mobile_toolbar.click_more_tab_mobile()
        self.mobile_toolbar.click_fattal_friends_club_tab()
        self.mobile_club_join_page.click_join_fattal_friends_button()

        # Step 1: Navigate to renewal page

        # Step 2: Agreement & Payment
        self.mobile_order_page.click_user_agreement_checkbox_by_label_id()
        sleep(15)
        self.take_stage_screenshot("payment_stage")
        self.fill_payment_details_from_config()

        # Step 3: Submit payment
        self.mobile_order_page.click_payment_submit_button()

        # Step 4: Verify price
        try:
            logging.info("Waiting for confirmation price element...")
            price_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.ID, "club-checkout-order-details-price"))
            )
            price_text = price_element.text.strip().replace('\u200f', '').replace('\xa0', ' ')
            logging.info(f"Found price text: '{price_text}'")

            expected_prices = ["150 ₪", "99 ₪"]
            assert any(expected in price_text for expected in expected_prices), \
                f"Price mismatch! Expected one of {expected_prices}, got '{price_text}'"

            logging.info("Final price matches one of the expected values.")
        except Exception as e:
            logging.error(f"Failed to assert price: {e}")
            raise
        self.mobile_toolbar.close_any_club_popup()
        # Step 5: Optional Confirmation Log
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        if self.confirmation_result:
            logging.info("ℹ️ Confirmation page loaded. Skipping order number check.")

            # 📸 Screenshot confirmation page for renewal
            self.confirmation_screenshot_path = self.take_confirmation_screenshot_renew_membership(self._testMethodName,
                                                                                                   "success")
            setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)
        else:
            logging.warning("⚠️ Confirmation result was empty — order number not available.")

    def test_mobile_booking_club_member_11night(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר חבר מועדון פעיל + הטבת לילה 11 מתנה"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_11NIGHT_ID"),
            "password": os.getenv("CLUB_11NIGHT_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)
    def test_mobile_booking_club_member_deals(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר עמוד דילים"
        user = {
            "id": os.getenv("CLUB_REGULAR_ID"),
            "password": os.getenv("CLUB_REGULAR_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # ✅ For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        self.mobile_toolbar.click_deals_and_packages_tab()
        self.mobile_deals_page.click_view_all_deals_link()
        self.mobile_deals_page.click_view_more_deal_button()
        self.mobile_deals_page.click_book_now_button()
        self.mobile_deals_page.click_continue_search_button_mobile()
        self.mobile_deals_page.click_continue_room_button()
        self.mobile_deals_page.click_mobile_search_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_deals_page.click_mobile_show_prices_button()
        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.click_user_agreement_checkbox()
        self.take_stage_screenshot("payment_stage")
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)
    def test_mobile_booking_anonymous_user_promo_code(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי ושימוש בפרומו קוד חבר (FHVR)"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_months_ahead(months_ahead=3, stay_length=3)

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.open_promo_code_input()
        self.mobile_main_page.enter_promo_code("FHVR")
        assert self.mobile_main_page.is_promo_code_applied("FHVR"), "Promo code was not correctly applied!"

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        #Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()
        #Step 6 : Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        #Order Details
        self.fill_guest_details(guest=self.default_guest)


        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_anonymous_fattal_employee_promo_code(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי ושימוש בפרומו קוד חבר (EMP23FA)"
        hotel_name = self.default_hotel_name
        user = {
            "id": os.getenv("EMPLOYEE_COUPON_ID"),
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
        }
        logging.info("Starting test: hotel search and booking flow")

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Promo code input
        self.mobile_main_page.open_promo_code_input()
        self.mobile_main_page.enter_promo_code("EMP23FA")
        assert self.mobile_main_page.is_promo_code_applied("EMP23FA"), "Promo code was not correctly applied!"

        # Step 5: Perform the search
        self.mobile_main_page.click_mobile_search_button()
        self.mobile_main_page.enter_id(user["id"])
        self.mobile_main_page.click_validation_button()

        # Step 6: Choose Room
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        # Step 7: Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")

        # ✅ Fill in guest details
        self.mobile_order_page.set_email(user["email"])
        self.mobile_order_page.set_phone(user["phone"])
        self.entered_id_number = user["id"]

        # ✅ Save for export/logging
        self.entered_email = user["email"]

        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)

        # Step 8: Fill payment iframe
        self.fill_payment_details_from_config()

        # Step 9: Submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_anonymous_europe(self):
        self.save_for_cancellation = False  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        hotel_name = self.default_hotel_name_europe
        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5: Choose Room and click it
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        # Step 6: Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        # Order Details
        self.fill_guest_details(guest=self.default_guest_europe)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()

        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_with_club_login_europe(self):
        self.save_for_cancellation = False  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        hotel_name = self.default_hotel_name_europe
        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_REGULAR_ID"),
            "password": os.getenv("CLUB_REGULAR_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")

        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5: Choose Room and click it
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        # Step 6: Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.set_first_name("Chen")
        self.mobile_order_page.set_last_name("Test")
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()

        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)

    def test_mobile_booking_with_club_login_11night_europe(self):
        self.save_for_cancellation = False  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        hotel_name = self.default_hotel_name_europe
        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        user = {
            "id": os.getenv("CLUB_11NIGHT_ID_EUROPE"),
            "password": os.getenv("CLUB_11NIGHT_PASSWORD_EUROPE")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")

        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()
        self.handle_no_results_and_click_suggestion()
        # Step 5: Choose Room and click it
        self.mobile_search_page.click_show_prices_regional()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_regional()

        # Step 6: Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")
        self.mobile_order_page.set_first_name("Chen")
        self.mobile_order_page.set_last_name("Test")
        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)
    def test_mobile_booking_anonymous_user_login_at_checkout(self):
        self.save_for_cancellation = True  # Enable save-for-cancel feature

        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי וחיבור במסך תשלום"
        hotel_name = self.default_hotel_name
        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.close_war_popup()
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()
        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        #Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.take_stage_screenshot("room_selection")
        self.mobile_search_page.click_book_room_button()

        #Step 6 : Order Page
        #self.mobile_order_page.click_room_selection_summary()

        self.mobile_order_page.wait_until_personal_form_ready()

        #Order Details
        self.take_stage_screenshot("payment_stage")
        user = {
            "id": os.getenv("CLUB_REGULAR_ID"),
            "password": os.getenv("CLUB_REGULAR_PASSWORD")
        }
        try:
            self.mobile_toolbar.open_login_menu()
            self.mobile_toolbar.click_login_with_email_button()
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_id_number = user["id"]
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        self.mobile_order_page.click_user_agreement_checkbox()
        sleep(15)
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9: Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.",
                         self.soft_assert_errors)
        if self.soft_assert_errors:
            logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

        # Step 10: Always take a screenshot of the confirmation screen
        self.confirmation_screenshot_path = self.take_confirmation_screenshot(self._testMethodName, "success")
        setattr(self, "screenshot_confirmation", self.confirmation_screenshot_path)
    # def test_mobile_booking_5_rooms_club_member(self):
    #     self.save_for_cancellation = True  # Enable save-for-cancel feature
    #
    #     self.soft_assert_errors = []
    #
    #     self.test_description = "בדיקת השלמת הזמנה משתמש מחובר חבר מועדון פעיל"
    #     hotel_name = self.default_hotel_name
    #
    #     logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")
    #
    #     # Step 0: Club Login
    #     user = {
    #         "id": os.getenv("CLUB_REGULAR_ID"),
    #         "password": os.getenv("CLUB_REGULAR_PASSWORD")
    #     }
    #     try:
    #         self.mobile_toolbar.open_login_menu()
    #         self.mobile_toolbar.user_id_input().send_keys(user["id"])
    #         self.mobile_toolbar.user_password_input().send_keys(user["password"])
    #         self.mobile_toolbar.click_login_button()
    #         self.mobile_toolbar.close_post_login_popup()
    #         logging.info("Logged in successfully.")
    #     except Exception as e:
    #         logging.warning(f"Login failed or already logged in: {e}")
    #     # For report logging only — because form fields are autofilled
    #     self.entered_id_number = user["id"]
    #     self.entered_first_name = "Club"
    #     self.entered_last_name = "User"
    #     # Step 1: City selectionX
    #     self.mobile_main_page.click_mobile_hotel_search_input()
    #     self.mobile_main_page.set_city_mobile(hotel_name)
    #     self.mobile_main_page.click_first_suggested_hotel()
    #
    #     # Step 2: Date picker
    #     self.mobile_main_page.click_mobile_date_picker()
    #     self.mobile_main_page.select_date_range_two_months_ahead()
    #
    #     # Step 3: Room selection
    #     self.mobile_main_page.click_mobile_room_selection()
    #     self.mobile_main_page.set_five_room_occupants()
    #     self.mobile_main_page.click_room_continue_button()
    #
    #     # Step 4: Perform the search
    #     self.mobile_main_page.click_mobile_search_button()
    #
    #     # Step 5 : Choose Room and click it
    #     #1
    #     # Step 5: Choose Room and click it (robust with fallback)
    #     self.mobile_search_page.click_show_then_book_room_with_fallback()
    #
    #     #2
    #     # Step 5: Choose Room and click it (robust with fallback)
    #     self.mobile_search_page.click_show_then_book_room_with_fallback()
    #
    #     #3
    #     # Step 5: Choose Room and click it (robust with fallback)
    #     self.mobile_search_page.click_show_then_book_room_with_fallback()
    #
    #     #4
    #     # Step 5: Choose Room and click it (robust with fallback)
    #     self.mobile_search_page.click_show_then_book_room_with_fallback()
    #
    #     #5
    #     # Step 5: Choose Room and click it (robust with fallback)
    #     self.mobile_search_page.click_show_then_book_room_with_fallback()
    #     self.take_stage_screenshot("room_selection")
    #
    #     # Step 6 : Order Page (for club, skip email + id)
    #     self.mobile_order_page.wait_until_personal_form_ready()
    #     self.take_stage_screenshot("payment_stage")
    #     self.mobile_order_page.click_user_agreement_checkbox()
    #     sleep(15)
    #     #test
    #     # Step 7: Fill the iframe using config.json
    #     self.fill_payment_details_from_config()
    #
    #     # Step 8: Click submit inside iframe (already inside from step 7)
    #     self.mobile_order_page.click_payment_submit_button()
    #     #Step 9 : Confirm and Assert
    #     self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
    #     self.soft_assert(self.confirmation_result.get("order_number"), "Booking failed — no order number found.", self.soft_assert_errors)
    #     if self.soft_assert_errors:
    #        logging.error("Soft assertions encountered:\n" + "\n".join(self.soft_assert_errors))

    def tearDown(self):
        if self.driver:
            try:
                has_failed = False
                outcome = getattr(self, "_outcome", None)
                if outcome:
                    result = outcome.result if hasattr(outcome, "result") else outcome
                    for method_name, exc_info in (getattr(result, "failures", []) + getattr(result, "errors", [])):
                        if method_name.id().endswith(self._testMethodName):
                            has_failed = True
                            break

                self.post_test_logging_mobile(self._test_result_for_teardown)

                # ✅ Save order for cancellation automatically if enabled
                if getattr(self, "save_for_cancellation", False):
                    order_number = getattr(self, "confirmation_result", {}).get("order_number")
                    if order_number:
                        self.save_order_for_cancellation(order_number)

            except Exception as e:
                logging.warning(f"Logging failed during tearDown: {e}")
            finally:
                logging.info("Waiting 2 seconds before closing browser...")
                sleep(2)
                try:
                    self.driver.quit()
                except Exception as e:
                    logging.warning(f"Browser quit failed: {e}")

    if __name__ == "__main__":
        import unittest
        unittest.main()










