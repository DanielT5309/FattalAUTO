import io
import logging
import sys
import time
import traceback
from openpyxl.styles import Font, PatternFill
from time import sleep
from unittest import TestCase
from selenium import webdriver
import random
from Mobile_Fattal_Pages.Mobile_Fattal_Flight_Page import FattalFlightPageMobile
from Mobile_Fattal_Pages.Mobile_Toolbar_Fattal import FattalMobileToolBar
from Mobile_Fattal_Pages.Mobile_Fattal_Main_Page import FattalMainPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Search_Result_Page import FattalSearchResultPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Order_Page import FattalOrderPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_ConfirmPage import FattalMobileConfirmPage
from Mobile_Fattal_Pages.Mobile_Fattal_Deals_Packages import FattalDealsPageMobile
from Mobile_Fattal_Pages.Mobile_Fattal_Customer_Contact_Page import FattalMobileCustomerSupport
from Mobile_Fattal_Pages.Mobile_Fattal_Join_Club_Page import FattalMobileClubJoinPage
import platform
from datetime import datetime
from faker import Faker
import os
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
class FattalTests(TestCase):
    def setUp(self):
        "Setup"
        load_dotenv()
        self.log_stream = io.StringIO()
        # Reset logging configuration
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(self.log_stream),
                logging.StreamHandler(sys.stdout)
            ]
        )

        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        # Use a valid built-in Chrome emulation device (Pixel 2)
        mobile_emulation = {
            "deviceName": "Pixel 2"
        }

        options = webdriver.ChromeOptions()
        options.add_experimental_option("mobileEmulation", mobile_emulation)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--disable-infobars")

        self.test_start_time = datetime.now()
        self.driver = webdriver.Chrome(options=options)
        self.driver.implicitly_wait(10)

        # Optional — force touch emulation for some mobile interactions
        self.driver.execute_cdp_cmd("Emulation.setTouchEmulationEnabled", {
            "enabled": True,
            "configuration": "mobile"
        })
        self.driver.set_window_rect(x=540, y=0, width=420, height=800)
        active_key = os.getenv("ENV_ACTIVE")
        self.driver.get(active_key)
        logging.info(f"Opened environment URL: {active_key}")
        # Load defaults from .env
        self.default_guest = {
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
            "first_name": os.getenv("DEFAULT_FIRST_NAME"),
            "last_name": os.getenv("DEFAULT_LAST_NAME")
        }
        self.default_guest_europe = {
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
            "first_name": os.getenv("DEFAULT_FIRST_NAME_ENGLISH"),
            "last_name": os.getenv("DEFAULT_LAST_NAME_ENGLISH")
        }
        self.default_hotel_name = os.getenv("DEFAULT_HOTEL_NAME")
        self.default_hotel_name_europe = os.getenv("DEFAULT_HOTEL_NAME_EUROPE")
        self.payment_card = {
            "card_number": os.getenv("PAYMENT_CARD_NUMBER"),
            "cardholder_name": os.getenv("PAYMENT_CARDHOLDER_NAME"),
            "expiry_month": os.getenv("PAYMENT_EXPIRY_MONTH"),
            "expiry_year": os.getenv("PAYMENT_EXPIRY_YEAR"),
            "cvv": os.getenv("PAYMENT_CVV"),
            "id_number": os.getenv("PAYMENT_ID_NUMBER")
        }

        # Page object initialization
        self.mobile_main_page = FattalMainPageMobile(self.driver)
        self.mobile_search_page = FattalSearchResultPageMobile(self.driver)
        self.mobile_order_page = FattalOrderPageMobile(self.driver)
        self.mobile_toolbar = FattalMobileToolBar(self.driver)
        self.mobile_confirm = FattalMobileConfirmPage(self.driver)
        self.mobile_deals_page = FattalDealsPageMobile(self.driver)
        self.mobile_flight_page = FattalFlightPageMobile(self.driver)
        self.mobile_customer_support = FattalMobileCustomerSupport(self.driver)
        self.mobile_club_join_page = FattalMobileClubJoinPage(self.driver)
    def post_test_logging(self, result):
        test_method = self._testMethodName
        has_failed = False
        error_msg = ""
        screenshot_path = ""
        log_file_path = ""
        duration = (datetime.now() - self.test_start_time).total_seconds()

        try:
            browser = self.driver.capabilities['browserName'] if self.driver else "unknown"
        except Exception as e:
            logging.warning(f"Could not get browser info: {e}")
            browser = "unknown"

        os_name = platform.system()

        # Save logs
        log_summary = self.log_stream.getvalue()
        logs_dir = os.path.join(self.base_dir, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        log_file_path = os.path.join(logs_dir, f"{test_method}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        try:
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(log_summary)
            print(f"[LOG DEBUG] Written log to {log_file_path}")
        except Exception as e:
            print(f"[LOG DEBUG] Failed to write log: {e}")

        # Detect real test failure using unittest outcome
        try:
            outcome = getattr(result, 'result', result)
            for failed_test, exc_info in outcome.failures + outcome.errors:
                if self._testMethodName in str(failed_test):
                    has_failed = True
                    if len(exc_info) == 3:
                        exc_type, exc_value, tb = exc_info
                        error_msg = "".join(traceback.format_exception(exc_type, exc_value, tb))
                    else:
                        error_msg = f" Unrecognized exception format: {exc_info}"
                    break
        except Exception as e:
            logging.warning(f"Failed to analyze test outcome: {e}")

        # Capture confirmation metadata if available
        confirmation = getattr(self, "confirmation_result", {})
        order_number = confirmation.get("order_number", "")
        confirmed_email = confirmation.get("email", "")
        id_number = getattr(self, 'entered_id_number', '')
        confirmation_screenshot = confirmation.get("screenshot_path", "")

        # Create one unified screenshot depending on test result
        try:
            if self.driver:
                status_label = "FAIL" if has_failed else ("PASS" if order_number else "FAIL")
                screenshot_path = self.take_confirmation_screenshot(test_method, status_label)
                confirmation_screenshot = screenshot_path
        except Exception as e:
            logging.warning(f"Could not take confirmation screenshot: {e}")

        # Final test info dict
        test_info = {
            "name": test_method,
            "description": getattr(self, "test_description", "No description provided"),
            "status": "FAILED" if has_failed else "PASSED",
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "duration": f"{duration:.2f}s",
            "browser": browser,
            "id_number": getattr(self, 'entered_id_number', ''),
            "os": os_name,
            "full_name": f"{getattr(self, 'entered_first_name', '')} {getattr(self, 'entered_last_name', '')}".strip(),
            "email": confirmed_email or getattr(self, 'entered_email', ''),
            "order_number": order_number,
            "screenshot": confirmation_screenshot,
            "log": log_file_path,
            "error": error_msg
        }

        # Persist results
        self.save_to_excel(test_info)
        self.save_to_pdf(test_info)

    def save_to_excel(self, info: dict):
        parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(parent_folder, "TestResults.xlsx")

        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results"
                ws.append([
                    "Test Name", "Description", "Status", "Timestamp", "Duration",
                    "Browser", "OS", "Full Name", "Email", "Order Number", "ID Number",
                    "Screenshot", "Log File"
                ])
            else:
                wb = load_workbook(filename)
                ws = wb.active

            status = "FAILED" if info.get("error") else "PASSED"

            row = [
                info.get("name", ""),
                info.get("description", ""),
                status,
                info.get("timestamp", ""),
                info.get("duration", ""),
                info.get("browser", ""),
                info.get("os", ""),
                info.get("full_name", ""),
                info.get("email", ""),
                info.get("order_number", ""),
                info.get("id_number", ""),
                info.get("screenshot", ""),
                info.get("log", "")
            ]
            ws.append(row)
            row_num = ws.max_row

            # 🎨 Add color to row based on pass/fail
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row_fill = red_fill if status == "FAILED" else green_fill

            for col_idx in range(1, 14):
                ws.cell(row=row_num, column=col_idx).fill = row_fill

            # 📎 Add screenshot hyperlink
            screenshot_path = info.get("screenshot", "")
            if screenshot_path and os.path.exists(screenshot_path):
                cell = ws.cell(row=row_num, column=12)
                cell.hyperlink = f"file:///{screenshot_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # 📎 Add log file hyperlink
            log_path = info.get("log", "")
            if log_path and os.path.exists(log_path):
                cell = ws.cell(row=row_num, column=13)
                cell.hyperlink = f"file:///{log_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # ↔️ Auto-fit column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.warning(f"Column resize error: {e}")
                ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(filename)
            logging.info(f"Excel file saved at: {filename}")

        except PermissionError as e:
            logging.warning(f"Excel file is open or locked: {e}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")

    def save_to_pdf(self, info: dict):
        pdf_dir = "reports"
        os.makedirs(pdf_dir, exist_ok=True)
        filename = f"{pdf_dir}/{info['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        c = canvas.Canvas(filename, pagesize=A4)

        width, height = A4
        y = height - 72

        c.setFont("Helvetica-Bold", 18)
        c.drawString(72, y, "Test Execution Report")
        y -= 40

        c.setFont("Helvetica", 12)
        lines = [
            f"Test Name: {info.get('name')}",
            f"Description: {info.get('description')}",
            f"Status: {info.get('status')}",
            f"Timestamp: {info.get('timestamp')}",
            f"Browser: {info.get('browser')}",
            f"OS: {info.get('os')}",
        ]
        if info.get("error"):
            lines.append(f"Error: {info.get('error')}")

        for line in lines:
            c.drawString(72, y, line)
            y -= 20

        if info.get("screenshot") and os.path.exists(info["screenshot"]):
            c.drawString(72, y, "Screenshot:")
            y -= 250
            c.drawImage(info["screenshot"], 72, y, width=5 * inch, preserveAspectRatio=True, mask='auto')

        c.showPage()
        c.save()
        print(f" PDF saved: {filename}")

    def fill_guest_details(self, guest=None, email=None, phone=None, first_name=None, last_name=None):
        """
        Fills guest details into the order page.
        You can either pass a full `guest` dictionary (with keys: email, phone, first_name, last_name)
        OR use individual arguments to override defaults.
        """
        if guest:
            email = guest.get("email", email)
            phone = guest.get("phone", phone)
            first_name = guest.get("first_name", first_name)
            last_name = guest.get("last_name", last_name)

        # Defaults fallback
        email = email or "chenttedgui@gmail.com"
        phone = phone or "0544531600"
        first_name = first_name or "חן"
        last_name = last_name or "טסט"

        self.mobile_order_page.set_email(email)
        self.mobile_order_page.set_phone(phone)
        self.mobile_order_page.set_first_name(first_name)
        self.mobile_order_page.set_last_name(last_name)

        # For logging/export
        self.entered_email = email
        self.entered_first_name = first_name
        self.entered_last_name = last_name

    def fill_payment_details_from_config(self):
        """
        Fills payment details using the credit card information from config.json
        """
        try:
            card_number = self.payment_card["card_number"]
            cardholder_name = self.payment_card["cardholder_name"]
            expiry_month = self.payment_card["expiry_month"]
            expiry_year = self.payment_card["expiry_year"]
            cvv = self.payment_card["cvv"]
            id_number = self.payment_card["id_number"]
            
            logging.info("Using credit card details from config.json")
            
            # Find and switch to the iframe
            iframe = self.driver.find_element(By.ID, "paymentIframe")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", iframe)
            self.driver.switch_to.frame(iframe)
            
            # Fill Card Number
            card_number_input = self.driver.find_element(By.ID, "credit_card_number_input")
            card_number_input.clear()
            card_number_input.send_keys(card_number)
            
            # Fill Cardholder Name
            cardholder_input = self.driver.find_element(By.ID, "card_holder_name_input")
            cardholder_input.clear()
            cardholder_input.send_keys(cardholder_name)
            
            # Select Expiry Month
            month_select = Select(self.driver.find_element(By.ID, "date_month_input"))
            month_select.select_by_visible_text(expiry_month)
            
            # Select Expiry Year
            year_select = Select(self.driver.find_element(By.ID, "date_year_input"))
            year_select.select_by_visible_text(expiry_year)
            
            # Fill CVV
            cvv_input = self.driver.find_element(By.ID, "cvv_input")
            cvv_input.clear()
            cvv_input.send_keys(cvv)
            
            # Fill ID Number
            id_input = self.driver.find_element(By.ID, "id_number_input")
            id_input.clear()
            id_input.send_keys(id_number)
            
            logging.info("Credit card details from config applied successfully")
            
        except Exception as e:
            logging.error(f"Failed to apply credit card details from config: {e}")
            raise
        finally:
            # Switch back to default content
            self.driver.switch_to.default_content()

    def take_confirmation_screenshot(self, test_method, status):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"confirmation_{status}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )
        self.driver.save_screenshot(filename)
        logging.info(f"Screenshot saved at: {filename}")
        return filename

    def run(self, result=None):
        self._resultForDoCleanups = result
        return super().run(result)

    def take_screenshot(self, test_method_name):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        screenshot_dir = os.path.join(os.getcwd(), "Fattal_Tests", "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = f"{screenshot_dir}/{test_method_name}_{timestamp}.png"
        self.driver.save_screenshot(filename)
        logging.error(f" Screenshot taken: {filename}")

    def test_mobile_join_fattal_and_friends_form(self):
        self.test_description = "הצטפרות למועדון דרך טופס"
        logging.info("Starting test: Join Fattal Club")

        try:
            # Step 1: Prepare data
            guest = self.default_guest
            first_name = guest["first_name"]
            last_name = guest["last_name"]
            email = guest["email"]
            phone = guest["phone"]
            birthdate = "01-01-1990"
            password = "Aa123456"
            id_number = self.mobile_order_page.generate_israeli_id()
            self.entered_id_number = id_number  # ✅ Store for logging/export
            logging.info(f"Generated ID for club registration: {id_number}")

            # Step 2: Navigate to club join screen
            self.mobile_toolbar.click_more_tab_mobile()
            self.mobile_toolbar.click_fattal_friends_club_tab()
            self.mobile_club_join_page.click_join_fattal_friends_button()

            # Step 3: Fill form
            self.mobile_club_join_page.fill_join_fattal_club_form(
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                id_number=id_number,
                birthdate=birthdate,
                password=password
            )

            # Step 4: Assert values
            self.mobile_club_join_page.assert_input_value("שם פרטי", first_name)
            self.mobile_club_join_page.assert_input_value("שם משפחה", last_name)
            self.mobile_club_join_page.assert_input_value("כתובת דוא״ל", email)
            self.mobile_club_join_page.assert_input_value("מספר טלפון נייד", phone)
            self.mobile_club_join_page.assert_input_value("תאריך לידה", birthdate)
            self.mobile_club_join_page.assert_input_value("מספר תעודת זהות", id_number)
            self.mobile_club_join_page.assert_input_value("בחרו סיסמא", password)

            logging.info("Fattal Club form filled and validated successfully.")
            self.mobile_club_join_page.click_accept_terms_checkbox()
            # Assuming you've already navigated to the payment step
            self.fill_payment_details_from_config()
            # Step 8: Switch BACK into iframe to click submit
            self.mobile_order_page.click_payment_submit_button()
            # Step 9 : Confirm and Assert
            #self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
            #assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

        except Exception as e:
            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
            screenshot_path = os.path.join("Fattal_Tests", "Screenshots", f"join_club_test_fail_{timestamp}.png")
            self.driver.save_screenshot(screenshot_path)
            logging.exception(f"Join Fattal Club test failed. Screenshot saved: {screenshot_path}")
            raise

    def test_mobile_contact_form(self):
        self.test_description = "בדיקת טופס צור קשר"

        logging.info("Starting test: Customer Support Ticket")

        try:
            guest = self.default_guest
            first_name = guest["first_name"]
            last_name = guest["last_name"]
            phone = guest["phone"]
            email = guest["email"]
            id_number = self.mobile_order_page.generate_israeli_id()
            message = "בדיקה אוטומטית של צור קשר דרך המובייל"
            hotel_name = "הרודס בוטיק אילת"

            # Step 1: Open the contact form
            self.mobile_toolbar.click_more_tab_mobile()
            self.mobile_toolbar.click_contact_us_button_mobile()
            self.mobile_customer_support.click_send_us_inquiry_button()

            # Step 2: Select 'נושא' dropdown
            self.mobile_customer_support.select_dropdown_by_label("נושא")

            # Step 3: Fill form inputs
            self.mobile_customer_support.fill_basic_contact_fields(
                first_name, last_name, id_number, phone, email, message, accept_marketing=False
            )

            # Step 4: Select hotel name
            self.mobile_customer_support.select_dropdown_by_label("שם המלון", option_text=hotel_name)

            # Step 5: Verify form data
            self.mobile_customer_support.assert_form_data_matches_input(
                first_name=first_name,
                last_name=last_name,
                id_number=id_number,
                phone=phone,
                email=email,
                message=message
            )

            logging.info("Test completed successfully.")

        except Exception as e:
            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
            screenshot_dir = os.path.join(os.getcwd(), "Fattal_Tests", "Screenshots")
            os.makedirs(screenshot_dir, exist_ok=True)
            screenshot_path = os.path.join(screenshot_dir, f"test_failure_{timestamp}.png")
            self.driver.save_screenshot(screenshot_path)
            logging.exception(f"Test failed. Screenshot saved: {screenshot_path}")
            raise

    def test_mobile_booking_anonymous_user(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי"
        hotel_name = self.default_hotel_name
        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        #Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()
        #Step 6 : Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        #Order Details
        self.fill_guest_details(guest=self.default_guest)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  #  Save for logging/export

        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9 : Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_anonymous_join_fattal_and_friends(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי + הצטפרות למועדון"

        hotel_name = self.default_hotel_name

        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()
        # Step 6 : Order Page
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.wait_until_personal_form_ready()
        # Order Details
        self.fill_guest_details(guest=self.default_guest)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9 : Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_eilat_with_flight(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר עם מועדון פעיל + טיסות"
        hotel_name = "אילת, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        logging.info("Starting mobile booking test for Eilat including flights...")

        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        #self.mobile_main_page.select_date_range_two_months_ahead()
        # Step 2: Select exact date range instead of the dynamic one
        self.mobile_main_page.select_specific_date_range(checkin_day=1, checkout_day=5)
        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        # Step 4: Enable flight option (🆕 use the mobile version of the method!)
        self.mobile_main_page.select_flight_option_all_airports()
        self.mobile_main_page.click_mobile_search_button()
        # Step 5: Handle results
        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.mobile_search_page.click_book_room_regional()

        # Step 7: Select flight or continue
        self.mobile_flight_page.try_flight_options_by_time_of_day()

        # Step 8: Passenger form
        self.mobile_flight_page.fill_adult_passenger_details()
        self.mobile_flight_page.click_continue_button()

        # Step 9: Payment form
        self.mobile_order_page.wait_until_personal_form_ready()
        # Order Details
        self.fill_guest_details(guest=self.default_guest)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 10: Confirmation
        #self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        #assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."

    def test_mobile_booking_anonymous_region_eilat(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי דרך אזור מלונות אילת"
        hotel_name = "אילת, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()
        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        # Step 3: Search Vacation
        self.mobile_main_page.click_room_continue_button(),





        
        self.mobile_main_page.click_mobile_search_button()

        # Step 5: Handle results
        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.mobile_search_page.click_book_room_regional()
        # Step 6 : Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        # Order Details
        self.fill_guest_details(guest=self.default_guest)

        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9 : Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_fattal_gift3(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי + קופון 1 של פתאל גיפטס באילת"
        hotel_name = "אילת, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")

        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead(stay_length=5)
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.click_mobile_search_button()

        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.mobile_search_page.click_book_room_regional()

        self.mobile_order_page.wait_until_personal_form_ready()
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()

        # Apply all 3 gift codes
        raw_gifts = [os.getenv("GIFT1"), os.getenv("GIFT2"), os.getenv("GIFT3")]
        gifts = [code.strip() for code in raw_gifts if code and code.strip().isdigit()]
        failed_gifts = []

        for code in gifts:
            logging.info(f"Applying gift coupon: '{code}'")
            self.mobile_order_page.apply_checkout_coupon(code)
            sleep(5)
            if not self.mobile_order_page.is_coupon_applied_successfully(code):
                logging.error(f"Gift code '{code}' was not applied successfully!")
                failed_gifts.append(code)
            else:
                logging.info(f"Gift code '{code}' applied successfully.")

        assert not failed_gifts, f"The following gift codes failed to apply: {', '.join(failed_gifts)}"

        self.fill_payment_details_from_config()
        self.mobile_order_page.click_payment_submit_button()

        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_fattal_gift1(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי +3 קופונים של פתאל גיפטס באילת"
        hotel_name = "אילת, ישראל"
        random_id = self.mobile_order_page.generate_israeli_id()
        logging.info(f"Generated Israeli ID: {random_id}")

        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_region()
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead(stay_length=5)
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=0, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.click_mobile_search_button()

        self.mobile_search_page.click_book_room_button()
        self.mobile_search_page.click_show_prices_regional()
        self.mobile_search_page.click_book_room_regional()

        self.mobile_order_page.wait_until_personal_form_ready()
        self.fill_guest_details(guest=self.default_guest)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()

        gift_code = os.getenv("GIFT4", "").strip()
        assert gift_code and gift_code.isdigit(), "GIFT4 is missing or invalid in environment config"

        logging.info(f"Applying single gift coupon: '{gift_code}'")
        self.mobile_order_page.apply_checkout_coupon(gift_code)
        sleep(5)

        applied_successfully = self.mobile_order_page.is_coupon_applied_successfully(gift_code)
        assert applied_successfully, f"Gift coupon '{gift_code}' failed to apply."

        self.fill_payment_details_from_config()
        self.mobile_order_page.click_payment_submit_button()

        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_club_member_club_renew_expired(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר מועדון בסטטוס פג תוקף"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        try:
            self.mobile_toolbar.open_login_menu()
            user = {
                "id": os.getenv("CLUB_RENEW_ID"),
                "password": os.getenv("CLUB_RENEW_PASSWORD")
            }
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")# For report logging only — because form fields are autofilled
        self.entered_first_name = "Club"
        self.entered_last_name = "User"


        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.click_user_agreement_checkbox()

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        # self.mobile_order_page.click_payment_submit_button()
        # # Step 9 : Confirm and Assert
        # self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        # assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."

    def test_mobile_booking_club_member_club_renew_about_to_expire(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר מועדון בסטטוס עומד לפוג"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        try:
            self.mobile_toolbar.open_login_menu()
            user = {
                "id": os.getenv("CLUB_ABOUT_EXPIRE_ID"),
                "password": os.getenv("CLUB_ABOUT_EXPIRE_PASSWORD")
            }
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.mobile_order_page.click_join_club_checkbox()
        self.mobile_order_page.click_user_agreement_checkbox()

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        # self.mobile_order_page.click_payment_submit_button()
        # # Step 9 : Confirm and Assert
        # self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        # assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."

    def test_mobile_booking_club_member_11night(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר חבר מועדון פעיל + הטבת לילה 11 מתנה"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        try:
            self.mobile_toolbar.open_login_menu()
            user = {
                "id": os.getenv("CLUB_11NIGHT_ID"),
                "password": os.getenv("CLUB_11NIGHT_PASSWORD")
            }
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_first_name = "Club"
        self.entered_last_name = "User"
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.mobile_order_page.click_user_agreement_checkbox()

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # # Step 8: Click submit inside iframe (already inside from step 7)
        # self.mobile_order_page.click_payment_submit_button()
        # #Step 9 : Confirm and Assert
        # self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        # assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_club_member(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר חבר מועדון פעיל"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")

        # Step 0: Club Login
        try:
            self.mobile_toolbar.open_login_menu()
            user = {
                "id": os.getenv("CLUB_REGULAR_ID"),
                "password": os.getenv("CLUB_REGULAR_PASSWORD")
            }
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # For report logging only — because form fields are autofilled
        self.entered_first_name = "Club"
        self.entered_last_name = "User"
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        # Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()

        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.wait_until_personal_form_ready()
        self.mobile_order_page.click_user_agreement_checkbox()

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        self.mobile_order_page.click_payment_submit_button()
        #Step 9 : Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."

    def test_mobile_booking_club_member_deals(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר עמוד דילים"
        try:
            self.mobile_toolbar.open_login_menu()
            user = {
                "id": os.getenv("CLUB_REGULAR_ID"),
                "password": os.getenv("CLUB_REGULAR_PASSWORD")
            }
            self.mobile_toolbar.user_id_input().send_keys(user["id"])
            self.mobile_toolbar.user_password_input().send_keys(user["password"])
            self.mobile_toolbar.click_login_button()
            self.mobile_toolbar.close_post_login_popup()
            logging.info("Logged in successfully.")
        except Exception as e:
            logging.warning(f"Login failed or already logged in: {e}")
        # ✅ For report logging only — because form fields are autofilled
        self.entered_first_name = "Club"
        self.entered_last_name = "User"

        self.mobile_toolbar.click_deals_and_packages_tab()
        self.mobile_deals_page.click_view_all_deals_link()
        self.mobile_deals_page.click_view_more_deal_button()
        self.mobile_deals_page.click_book_now_button()
        self.mobile_deals_page.click_continue_search_button_mobile()
        self.mobile_deals_page.click_continue_room_button()
        self.mobile_deals_page.click_mobile_search_button()
        self.mobile_deals_page.click_mobile_show_prices_button()
        # Step 6 : Order Page (for club, skip email + id)
        self.mobile_order_page.click_user_agreement_checkbox()

        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Click submit inside iframe (already inside from step 7)
        #self.mobile_order_page.click_payment_submit_button()
        # Step 9 : Confirm and Assert
        #self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        #assert self.confirmation_result.get("order_number"), "❌ Booking failed — no order number found."

    def test_mobile_booking_anonymous_random_guest_details(self):
        self.test_description = " בדיקה אנונימית רנדומלית ללא סיום הזמנה"
        fake = Faker('he_IL')  # עברית!
        hotel_name = self.default_hotel_name

        logging.info("Starting FULL RANDOM anonymous booking test (MOBILE)")

        # Step 1: חיפוש מלון ודילים
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: תאריכים
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: הרכב אורחים
        adults = random.randint(1, 2)
        children = random.randint(0, 1)
        infants = random.randint(0, 1)
        logging.info(f"Guests: {adults} adults, {children} children, {infants} infants")
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=adults, children=children, infants=infants)
        self.mobile_main_page.click_room_continue_button()

        # Step 4: חיפוש
        self.mobile_main_page.click_mobile_search_button()

        # Step 5: בחר חדר
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()

        # Step 6: הזנת פרטים רנדומליים
        self.mobile_order_page.wait_until_personal_form_ready()
        random_email = fake.email()
        random_phone = fake.phone_number().replace("-", "").replace("+", "")
        random_first_name = fake.first_name()
        random_last_name = fake.last_name()
        random_id = str(fake.random_int(min=100000000, max=999999999))
        random_note = fake.sentence(nb_words=6)

        self.entered_email = random_email
        self.entered_first_name = random_first_name
        self.entered_last_name = random_last_name

        self.mobile_order_page.set_email(random_email)
        self.mobile_order_page.set_phone(random_phone)
        self.mobile_order_page.set_first_name(random_first_name)
        self.mobile_order_page.set_last_name(random_last_name)
        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export

        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: פרטי כרטיס רנדומליים (לא נשלח בפועל)
        self.mobile_order_page.fill_payment_iframe_mobile_random()

        # לא נלחץ על כפתור שליחה - TEST MODE ONLY
        logging.info("Full random booking simulated (NO SUBMIT)")

        self.confirmation_result = {
            "order_number": "",  # אין מספר הזמנה כי לא ביצענו שליחה
            "email": self.entered_email,
            "screenshot_path": ""
        }

    def test_mobile_booking_anonymous_user_promo_code(self):
        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי ושימוש בפרומו קוד חבר (FHVR)"
        hotel_name = self.default_hotel_name

        logging.info("Starting test: hotel search and booking flow")
        random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
        logging.info(f"Generated Israeli ID: {random_id}")
        # Step 1: City selection
        self.mobile_main_page.click_mobile_hotel_search_input()
        self.mobile_main_page.set_city_mobile(hotel_name)
        self.mobile_main_page.click_first_suggested_hotel()

        # Step 2: Date picker
        self.mobile_main_page.click_mobile_date_picker()
        self.mobile_main_page.select_date_range_two_months_ahead()

        # Step 3: Room selection
        self.mobile_main_page.click_mobile_room_selection()
        self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
        self.mobile_main_page.click_room_continue_button()
        self.mobile_main_page.open_promo_code_input()
        self.mobile_main_page.enter_promo_code("FHVR")
        assert self.mobile_main_page.is_promo_code_applied("FHVR"), "Promo code was not correctly applied!"

        # Step 4: Perform the search
        self.mobile_main_page.click_mobile_search_button()

        #Step 5 : Choose Room and click it
        self.mobile_search_page.click_show_prices_button()
        self.mobile_search_page.click_book_room_button()
        #Step 6 : Order Page
        self.mobile_order_page.wait_until_personal_form_ready()
        #Order Details
        self.fill_guest_details(guest=self.default_guest)


        self.mobile_order_page.set_id_number(random_id)
        self.entered_id_number = random_id  # Save for logging/export
        self.mobile_order_page.click_user_agreement_checkbox()
        # Step 7: Fill the iframe using config.json
        self.fill_payment_details_from_config()

        # Step 8: Switch BACK into iframe to click submit
        self.mobile_order_page.click_payment_submit_button()
        # Step 9 : Confirm and Assert
        self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
        assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."
    # def test_mobile_booking_anonymous_europe(self):
    #     hotel_name = self.default_hotel_name_europe
    #     logging.info("Starting test: hotel search and booking flow")
    #     random_id = self.mobile_order_page.generate_israeli_id()  # Generate a valid Israeli ID
    #     logging.info(f"Generated Israeli ID: {random_id}")
    #     # Step 1: City selection
    #     self.mobile_main_page.click_mobile_hotel_search_input()
    #     self.mobile_main_page.set_city_mobile(hotel_name)
    #     self.mobile_main_page.click_first_suggested_hotel()
    #
    #     # Step 2: Date picker
    #     self.mobile_main_page.click_mobile_date_picker()
    #     self.mobile_main_page.select_date_range_two_months_ahead()
    #
    #
    #     # Step 3: Room selection
    #     self.mobile_main_page.click_mobile_room_selection()
    #     self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
    #     self.mobile_main_page.click_room_continue_button()
    #
    #     # Step 4: Perform the search
    #     self.mobile_main_page.click_mobile_search_button()
    #
    #     #Step 5 : Choose Room and click it
    #     self.mobile_search_page.click_show_prices_regional()
    #     self.mobile_search_page.click_book_room_regional()
    #     #Step 6 : Order Page
    #     self.mobile_order_page.wait_until_personal_form_ready()
    #     #Order Details
    #     self.fill_guest_details(guest=self.default_guest_europe)
    #
    #     self.mobile_order_page.set_id_number(random_id)
    #     self.entered_id_number = random_id  #  Save for logging/export
    #     self.mobile_order_page.click_user_agreement_checkbox()
    #     # Step 7: Fill the iframe using config.json
    #     self.fill_payment_details_from_config()
    #
    #     # Step 8: Switch BACK into iframe to click submit
    #     #self.mobile_order_page.click_payment_submit_button()
    #     # Step 9 : Confirm and Assert
    #     #self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
    #     #assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."



    # def test_mobile_booking_with_club_login_europe(self):
    #     hotel_name = self.default_hotel_name_europe
    #     logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")
    #
    #     # Step 0: Club Login
    #     try:
    #         self.mobile_toolbar.open_login_menu()
    #         user = {
    #             "id": os.getenv("CLUB_REGULAR_ID"),
    #             "password": os.getenv("CLUB_REGULAR_PASSWORD")
    #         }
    #         self.mobile_toolbar.user_id_input().send_keys(user["id"])
    #         self.mobile_toolbar.user_password_input().send_keys(user["password"])
    #         self.mobile_toolbar.click_login_button()
    #         self.mobile_toolbar.close_post_login_popup()
    #         logging.info("Logged in successfully.")
    #     except Exception as e:
    #         logging.warning(f"Login failed or already logged in: {e}")
    #     # For report logging only — because form fields are autofilled
    #     self.entered_first_name = "Club"
    #     self.entered_last_name = "User"
    #     # Step 1: City selection
    #     self.mobile_main_page.click_mobile_hotel_search_input()
    #     self.mobile_main_page.set_city_mobile(hotel_name)
    #     self.mobile_main_page.click_first_suggested_hotel()
    #
    #     # Step 2: Date picker
    #     self.mobile_main_page.click_mobile_date_picker()
    #     self.mobile_main_page.select_date_range_two_months_ahead()
    #
    #     # Step 3: Room selection
    #     self.mobile_main_page.click_mobile_room_selection()
    #     self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
    #     self.mobile_main_page.click_room_continue_button()
    #
    #     # Step 4: Perform the search
    #     self.mobile_main_page.click_mobile_search_button()
    #
    #     # Step 5 : Choose Room and click it
    #     self.mobile_search_page.click_show_prices_regional()
    #     self.mobile_search_page.click_book_room_regional()
    #
    #     # Step 6 : Order Page (for club, skip email + id)
    #     self.mobile_order_page.wait_until_personal_form_ready()
    #     self.mobile_order_page.set_first_name("Chen")
    #     self.mobile_order_page.set_last_name("Test")
    #     self.mobile_order_page.click_user_agreement_checkbox()
    #
    #     # Step 7: Fill the iframe using config.json
    #     self.fill_payment_details_from_config()
    #
    #     # Step 8: Click submit inside iframe (already inside from step 7)
    #     #self.mobile_order_page.click_payment_submit_button()
    #     #Step 9 : Confirm and Assert
    #     #self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
    #     #assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."


    # def test_mobile_booking_with_club_login_11night_europe(self):
    #     hotel_name = self.default_hotel_name_europe
    #     logging.info("Starting test: CLUB user hotel search and booking flow (mobile)")
    #
    #     # Step 0: Club Login
    #     try:
    #         self.mobile_toolbar.open_login_menu()
    #         user = {
    #             "id": os.getenv("CLUB_11NIGHT_ID"),
    #             "password": os.getenv("CLUB_11NIGHT_PASSWORD")
    #         }
    #         self.mobile_toolbar.user_id_input().send_keys(user["id"])
    #         self.mobile_toolbar.user_password_input().send_keys(user["password"])
    #         self.mobile_toolbar.click_login_button()
    #         self.mobile_toolbar.close_post_login_popup()
    #         logging.info("Logged in successfully.")
    #     except Exception as e:
    #         logging.warning(f"Login failed or already logged in: {e}")
    #     # For report logging only — because form fields are autofilled
    #     self.entered_first_name = "Club"
    #     self.entered_last_name = "User"
    #     # Step 1: City selection
    #     self.mobile_main_page.click_mobile_hotel_search_input()
    #     self.mobile_main_page.set_city_mobile(hotel_name)
    #     self.mobile_main_page.click_first_suggested_hotel()
    #
    #     # Step 2: Date picker
    #     self.mobile_main_page.click_mobile_date_picker()
    #     self.mobile_main_page.select_date_range_two_months_ahead()
    #
    #     # Step 3: Room selection
    #     self.mobile_main_page.click_mobile_room_selection()
    #     self.mobile_main_page.set_mobile_room_occupants(adults=2, children=1, infants=0)
    #     self.mobile_main_page.click_room_continue_button()
    #
    #     # Step 4: Perform the search
    #     self.mobile_main_page.click_mobile_search_button()
    #
    #     # Step 5 : Choose Room and click it
    #     self.mobile_search_page.click_book_room_button()
    #     self.mobile_search_page.click_show_prices_regional()
    #     self.mobile_search_page.click_book_room_regional()
    #
    #     # Step 6 : Order Page (for club, skip email + id)
    #     self.mobile_order_page.wait_until_personal_form_ready()
    #     self.mobile_order_page.set_first_name("Chen")
    #     self.mobile_order_page.set_last_name("Test")
    #     self.mobile_order_page.click_user_agreement_checkbox()
    #
    #     # Step 7: Fill the iframe using config.json
    #     self.fill_payment_details_from_config()
    #
    #     # # Step 8: Click submit inside iframe (already inside from step 7)
    #     # self.mobile_order_page.click_payment_submit_button()
    #     # #Step 9 : Confirm and Assert
    #     # self.confirmation_result = self.mobile_confirm.verify_confirmation_and_extract_order_mobile()
    #     # assert self.confirmation_result.get("order_number"), "Booking failed — no order number found."
    def tearDown(self):
        if self.driver:
            try:
                result = getattr(self, "_outcome", None)
                if result is None:
                    logging.warning("No result object found. Skipping post_test_logging.")
                else:
                    self.post_test_logging(result)

                # Log runtime duration
                try:
                    duration = self.driver.execute_script("""
                        const [start] = window.performance.getEntriesByName('selenium-start');
                        return Date.now() - start.startTime;
                    """)
                    logging.info(f" Test runtime: {int(duration)}ms")
                except Exception as e:
                    logging.warning(f" Could not get test runtime: {e}")

            except Exception as e:
                logging.warning(f"Logging failed during tearDown: {e}")

            logging.info("Waiting 2 seconds before closing browser...")
            sleep(2)  # Pause while browser is still open

            logging.info("Closing browser (tearDown).")
            try:
                self.driver.quit()
            except Exception as e:
                logging.warning(f"Browser quit failed: {e}")

    if __name__ == "__main__":
        import unittest
        unittest.main()











