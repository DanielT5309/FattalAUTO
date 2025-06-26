import json
import traceback
import unittest
from selenium import webdriver
from openpyxl.styles import  PatternFill
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Fattal_Pages.Fattal_Main_Page import FattalMainPage
from Fattal_Pages.Fattal_Tool_Bar import FattalToolBar
from Fattal_Pages.Fattal_Search_Result import FattalSearchResultPage
from Fattal_Pages.Fattal_Order_Page_ISR import FattalOrderPage
from Fattal_Pages.Fattal_Flight_OrderPage import FattalFlightOrderPage
from Fattal_Pages.Fattal_Confirmation_Page import FattalConfirmPage
import logging
import platform
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import io
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys
from dotenv import load_dotenv
from time import sleep
HOTEL_NAME_TO_ID = {
    "לאונרדו נגב, באר שבע": "10048",
    "לאונרדו פלאזה אילת": "10038"
}


def save_order_for_cancellation(master_id, hotel_id, filepath="orders_to_cancel.json"):
    import json
    orders = []
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            try:
                orders = json.load(f)
            except json.JSONDecodeError:
                pass

    orders.append({
        "masterID": master_id,
        "hotelID": hotel_id
    })

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(orders, f, ensure_ascii=False, indent=4)
class FattalDesktopTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.run_folder, cls.run_id = cls.get_static_run_folder()
    def setUp(self):
        self.run_folder = self.__class__.run_folder
        load_dotenv()

        options = webdriver.ChromeOptions()
        options.add_argument("--force-device-scale-factor=0.75")
        # options.add_argument("--headless")
        # options.add_argument("--no-sandbox")
        # options.add_argument("--disable-dev-shm-usage")

        self.test_start_time = datetime.now()
        self.log_stream = io.StringIO()

        # Base dir for logs/screenshots/etc.
        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        # Load default guest from .env
        self.default_guest = {
            "email": os.getenv("DEFAULT_EMAIL"),
            "phone": os.getenv("DEFAULT_PHONE"),
            "first_name": os.getenv("DEFAULT_FIRST_NAME"),
            "last_name": os.getenv("DEFAULT_LAST_NAME")
        }

        self.payment_card = {
            "card_number": os.getenv("PAYMENT_CARD_NUMBER"),
            "cardholder_name": os.getenv("PAYMENT_CARDHOLDER_NAME"),
            "expiry_month": os.getenv("PAYMENT_EXPIRY_MONTH"),
            "expiry_year": os.getenv("PAYMENT_EXPIRY_YEAR"),
            "cvv": os.getenv("PAYMENT_CVV"),
            "id_number": os.getenv("PAYMENT_ID_NUMBER")
        }

        # Reset logging
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(self.log_stream),
                logging.StreamHandler(sys.stdout)
            ]
        )

        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")

        self.driver = webdriver.Chrome(options=options)
        active_key = os.getenv("ENV_ACTIVE")
        self.driver.get(active_key)
        logging.info(f"Opened environment URL: {active_key}")
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.default_hotel_name = os.getenv("DEFAULT_HOTEL_NAME")

        # Page object setup
        self.main_page = FattalMainPage(self.driver)
        self.toolbar = FattalToolBar(self.driver)
        self.search_result = FattalSearchResultPage(self.driver)
        self.order_page = FattalOrderPage(self.driver)
        self.flight_page = FattalFlightOrderPage(self.driver)
        self.confirm_page = FattalConfirmPage(self.driver)

        # Start performance marker
        self.driver.execute_script("window.performance.mark('selenium-start')")
    def save_test_result_to_run_json(self, info: dict, run_folder: str):
        # 🔐 Ensure the folder exists
        os.makedirs(run_folder, exist_ok=True)

        json_path = os.path.join(run_folder, "run_data.json")

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:  # type: TextIO
                data = json.load(f)
        else:
            data = []

        data.append(info)

        with open(json_path, "w", encoding="utf-8") as f:  # type: TextIO
            json.dump(data, f, indent=2, ensure_ascii=False)

        logging.info(f"📄 Saved test result to: {json_path}")
    def soft_assert(self, condition, msg, errors_list):
        try:
            assert condition, msg
        except AssertionError as e:
            errors_list.append(str(e))

    @classmethod
    def get_static_run_folder(cls):
        base_path = os.path.join(os.path.dirname(__file__), 'html_reports', 'runs')
        os.makedirs(base_path, exist_ok=True)
        now = datetime.now().strftime("run_%Y-%m-%d_%H-%M-%S")
        run_folder = os.path.join(base_path, now)
        os.makedirs(run_folder, exist_ok=True)
        return run_folder, now

    @staticmethod
    def get_current_run_folder(cls=None):
        base_path = os.path.join(os.path.dirname(__file__), 'html_reports', 'runs')
        os.makedirs(base_path, exist_ok=True)
        now = datetime.now().strftime("run_%Y-%m-%d_%H-%M-%S")
        run_folder = os.path.join(base_path, now)
        os.makedirs(run_folder, exist_ok=True)
        return run_folder, now

    @staticmethod
    @staticmethod
    def generate_dashboard_html(runs_base_dir: str):
        import os
        import json
        from datetime import datetime
        import logging

        dashboard_path = os.path.join(runs_base_dir, "..", "dashboard.html")
        run_dirs = sorted(
            [d for d in os.listdir(runs_base_dir) if d.startswith("run_")],
            reverse=True
        )

        run_data_js_entries = []
        select_html_entries = []

        for i, run_dir in enumerate(run_dirs):
            json_path = os.path.join(runs_base_dir, run_dir, "run_data.json")
            if not os.path.exists(json_path):
                continue

            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                run_data_js_entries.append(f'"{run_dir}": {json.dumps(data, ensure_ascii=False)}')

                total = len(data)
                failed = sum(1 for t in data if t.get("status") == "FAILED")
                passed = total - failed
                mobile = sum(1 for t in data if t.get("test_type") == "mobile")
                desktop = sum(1 for t in data if t.get("test_type") == "desktop")

                run_timestamp_str = run_dir.replace("run_", "")  # e.g., '2025-06-17_14-33-12'
                run_datetime = datetime.strptime(run_timestamp_str, "%Y-%m-%d_%H-%M-%S")
                day_label = run_datetime.strftime("%A")
                datetime_label = run_datetime.strftime("%Y-%m-%d %H:%M:%S")

                label = f"{day_label} {datetime_label} | {total} tests | ✅ {passed} | ❌ {failed}"
                if mobile:
                    label += f" | 📱 {mobile}"
                if desktop:
                    label += f" | 💻 {desktop}"

                selected = "selected" if i == 0 else ""
                select_html_entries.append(f'<option value="{run_dir}" {selected}>{label}</option>')

        run_data_js = "{\n" + ",\n".join(run_data_js_entries) + "\n}"
        select_html = "\n".join(select_html_entries)

        script = f"""
            <script>
            const runData = {run_data_js};

            function populateRun(runId) {{
                const container = document.getElementById("results");
                container.innerHTML = "";
                const showPassed = document.getElementById("filterPassed").checked;
                const showFailed = document.getElementById("filterFailed").checked;
                const data = runData[runId] || [];
                data.forEach(test => {{
                    if ((test.status === "PASSED" && !showPassed) || (test.status === "FAILED" && !showFailed)) return;

                    const div = document.createElement("div");
                    div.classList.add("test-entry");
                    const screenshots = ["room_selection", "payment_stage", test.status === "FAILED" ? "error_screenshot" : "confirmation_screenshot"]
                      .map(label => {{
                          const path = test[label] || "";
                          if (!path) return "";
                          const short = path.split(/[/\\\\]/).pop();
                          return `<div><strong>${{label}}:</strong><br><img src="../Screenshots/${{short}}" style="max-height:120px;cursor:pointer;" onclick="openModal(this.src)" /></div>`;
                      }}).join("");

                    // ---- Log path robust logic ----
                    const logPath = test.log || "";
                    let logHref = "#";
                    if (logPath) {{
                      let parts = logPath.replace(/\\\\/g, '/').split('/');
                      if (parts.length >= 2) {{
                        logHref = "../" + parts.slice(-2).join('/');
                      }} else {{
                        logHref = logPath;
                      }}
                    }}
                    // --------------------------------

                    div.innerHTML = `
                      <h3>${{test.name}} — <span style="color:${{test.status === 'PASSED' ? 'green' : 'red'}}">${{test.status}}</span></h3>
                      <p><strong>Description:</strong> ${{test.description || "—"}}</p>
                      <p><strong>Timestamp:</strong> ${{test.timestamp}} | <strong>Duration:</strong> ${{test.duration}}</p>
                      <p><strong>Guest:</strong> ${{test.full_name}} | <strong>Email:</strong> ${{test.email}}</p>
                      <p><strong>Order #:</strong> ${{test.order_number}} | <strong>ID:</strong> ${{test.id_number}}</p>
                      <p><strong>Log:</strong> <a href="${{logHref}}" target="_blank">${{logPath.split(/[\\\\/]/).pop()}}</a></p>
                      ${{test.error ? `<p style='color:red'><strong>Error:</strong> ${{test.error}}</p>` : ""}}
                      <div class="screenshot-grid">${{screenshots}}</div>
                      <hr>`;
                    container.appendChild(div);
                }});
            }}

            function openModal(src) {{
                const modal = document.getElementById("screenshotModal");
                const modalImg = document.getElementById("modalImage");
                modal.style.display = "block";
                modalImg.src = src;
            }}

            function closeModal() {{
                document.getElementById("screenshotModal").style.display = "none";
            }}
            </script>
            """

        html = f"""<!DOCTYPE html>
        <html>
        <head>
          <meta charset="UTF-8">
          <title>🧪 Fattal Run Selector Dashboard</title>
          <style>
            body {{
              font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
              background-color: #f9f9f9;
              margin: 20px;
              color: #333;
            }}
            .dashboard-header {{
              display: flex;
              align-items: center;
              justify-content: space-between;
              margin-bottom: 10px;
              gap: 24px;
            }}
            .dashboard-header h1 {{
              margin: 0;
              font-size: 1.8em;
              display: flex;
              align-items: center;
              white-space: nowrap;
            }}
            .dashboard-header h1::before {{
              content: '🧪';
              margin-right: 10px;
            }}
            .header-logo {{
              height: 60px;
              max-width: 200px;
              object-fit: contain;
              border-radius: 8px;
              background: #fff;
              padding: 6px 12px;
              box-shadow: 0 2px 8px rgba(0,0,0,0.04);
            }}
            @media (max-width: 700px) {{
              .dashboard-header {{
                flex-direction: column;
                align-items: flex-start;
                gap: 8px;
              }}
              .header-logo {{
                margin-left: 0;
                margin-top: 6px;
                height: 44px;
                max-width: 150px;
              }}
            }}
            select {{
              font-size: 14px;
              padding: 5px;
              margin-left: 10px;
            }}
            .test-entry {{
              background: #fff;
              border-radius: 6px;
              padding: 15px;
              margin-bottom: 20px;
              box-shadow: 0 1px 4px rgba(0,0,0,0.1);
              transition: background 0.3s ease;
            }}
            .test-entry:hover {{
              background: #f0f8ff;
            }}
            .screenshot-grid {{
              display: flex;
              gap: 12px;
              margin-top: 10px;
              flex-wrap: wrap;
            }}
            .screenshot-grid img {{
              border-radius: 4px;
              border: 1px solid #ccc;
            }}
            .modal {{
              display: none;
              position: fixed;
              z-index: 999;
              left: 0; top: 0; width: 100%; height: 100%;
              background-color: rgba(0,0,0,0.85);
            }}
            .modal-content {{
              margin: 5% auto;
              display: block;
              max-width: 90vw;
              max-height: 80vh;
            }}
            .close {{
              position: absolute;
              top: 15px;
              right: 35px;
              color: #fff;
              font-size: 40px;
              font-weight: bold;
              cursor: pointer;
            }}
          </style>
          {script}
        </head>
        <body onload="populateRun(document.getElementById('runSelect').value)">
          <div class="dashboard-header">
            <h1>Fattal Run Selector Dashboard</h1>
            <img src="https://d2nyvxq412w7ra.cloudfront.net/_fcb5f25e78.png" alt="Fattal Logo" class="header-logo" />
          </div>
          <label>Choose Run:
            <select id="runSelect" onchange="populateRun(this.value)">
              {select_html}
            </select>
          </label>
          <div style="margin-top: 10px;">
            <label><input type="checkbox" id="filterPassed" checked onchange="populateRun(document.getElementById('runSelect').value)"> Show Passed</label>
            <label><input type="checkbox" id="filterFailed" checked onchange="populateRun(document.getElementById('runSelect').value)"> Show Failed</label>
          </div>
          <div id="results" style="margin-top: 20px;"></div>
          <div id="screenshotModal" class="modal" onclick="closeModal()">
            <span class="close">&times;</span>
            <img class="modal-content" id="modalImage">
          </div>
        </body>
        </html>"""

        os.makedirs(os.path.dirname(dashboard_path), exist_ok=True)
        with open(dashboard_path, "w", encoding="utf-8") as f:
            f.write(html)

        logging.info(f"✅ Dashboard generated at: {dashboard_path}")

    def confirm_and_assert_order(self):
        # Ensure confirmation_result is assigned before checking it
        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        if not self.confirmation_result:
            logging.error("Failed to extract confirmation details.")
            self.fail("Booking confirmation details not found.")

        self.confirmation_result["id_number"] = self.entered_id_number

        # Soft assertion for order number
        self.soft_assert(
            condition=self.confirmation_result.get("order_number"),
            msg="❌ Booking failed — no order number found.",
            errors_list=self.soft_assert_errors
        )
        if self.soft_assert_errors:
            self.fail("Soft assertions failed:\n" + "\n".join(self.soft_assert_errors))

    def retry_flight_search_if_no_results(self, hotel_name):
        def no_results_displayed():
            try:
                msg = self.driver.find_element(By.ID, "search-page-no-search-results-title")
                return "לא נמצאו מלונות" in msg.text
            except:
                return False

        if no_results_displayed():
            logging.warning("No flight hotel results found. Retrying...")

            self.take_screenshot("flight_search_no_results_initial")

            # Retry flight hotel search
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.select_flight_option_all_airports()
            self.main_page.search_button()

            # Wait and verify results
            try:
                WebDriverWait(self.driver, 15).until_not(
                    EC.presence_of_element_located((By.ID, "search-page-no-search-results-title"))
                )
                logging.info("Flight results appeared after retry.")
            except TimeoutException:
                self.take_screenshot("flight_search_retry_failed")
                raise Exception("No flight results even after retry.")
        else:
            logging.info("Flight hotel results returned.")

    def post_test_logging(self, result):
        test_method = self._testMethodName
        has_failed = False
        error_msg = ""
        duration = (datetime.now() - self.test_start_time).total_seconds()
        screenshot_path = ""
        error_screenshot_path = ""
        confirmation_screenshot_path = ""
        log_file_path = ""

        # ── Browser Info ──
        try:
            browser = self.driver.capabilities.get("browserName", "unknown") if self.driver else "unknown"
        except Exception as e:
            logging.warning(f"Could not get browser info: {e}")
            browser = "unknown"

        os_name = platform.system()

        # ── Save Logs ──
        log_summary = self.log_stream.getvalue()
        logs_dir = os.path.join(self.base_dir, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        log_file_path = os.path.join(logs_dir, f"{test_method}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        try:
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(log_summary)
            logging.info(f"[LOG DEBUG] Written log to {log_file_path}")
        except Exception as e:
            logging.error(f"[LOG DEBUG] Failed to write log: {e}")

        # ── Detect Failure ──
        try:
            outcome = getattr(result, 'result', result)
            for failed_test, exc_info in outcome.failures + outcome.errors:
                if self._testMethodName in str(failed_test):
                    has_failed = True
                    if len(exc_info) == 3:
                        exc_type, exc_value, tb = exc_info
                        error_msg = "".join(traceback.format_exception(exc_type, exc_value, tb))
                    else:
                        error_msg = f"Unrecognized exception format: {exc_info}"
                    break
        except Exception as e:
            logging.warning(f"Failed to analyze test outcome: {e}")

        # ── Confirmation Data ──
        confirmation = getattr(self, "confirmation_result", {}) or {}
        order_number = confirmation.get("order_number", "")
        confirmed_email = confirmation.get("email", "") or getattr(self, 'entered_email', "")

        # ── Screenshots ──
        try:
            if self.driver:
                if has_failed:
                    error_screenshot_path = self.take_confirmation_screenshot(test_method, "FAIL")
                else:
                    confirmation_screenshot_path = self.take_confirmation_screenshot(test_method, "PASS")
        except Exception as e:
            logging.warning(f"Could not take screenshot: {e}")

        # ── Test Type Detection ──
        test_type = "desktop"
        try:
            user_agent = self.driver.execute_script("return navigator.userAgent")
            if "Mobile" in user_agent:
                test_type = "mobile"
        except Exception as e:
            logging.warning(f"Could not determine user agent: {e}")

        # ── Final Info Dictionary ──
        test_info = {
            "name": test_method,
            "description": getattr(self, "test_description", "No description provided"),
            "status": "FAILED" if has_failed else "PASSED",
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "duration": f"{duration:.2f}s",
            "browser": browser,
            "os": os_name,
            "full_name": f"{getattr(self, 'entered_first_name', '')} {getattr(self, 'entered_last_name', '')}".strip(),
            "email": confirmed_email,
            "order_number": order_number,
            "id_number": getattr(self, "entered_id_number", ""),
            "confirmation_screenshot": confirmation_screenshot_path,
            "error_screenshot": error_screenshot_path,
            "room_selection": getattr(self, "screenshot_room_selection", ""),
            "payment_stage": getattr(self, "screenshot_payment_stage", ""),
            "log": log_file_path,
            "error": error_msg,
            "test_type": test_type
        }

        # Log soft assertion errors if they exist
        if self.soft_assert_errors:
            logging.info(f"Soft Assertion Errors found: {len(self.soft_assert_errors)}")
            for error in self.soft_assert_errors:
                logging.info(f"Soft Assertion Failure: {error}")

        # ── Save Results ──
        self.save_to_excel(test_info)
        self.save_test_result_to_run_json(test_info, self.run_folder)

    def take_stage_screenshot(self, label: str):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"{label}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )
        try:
            self.driver.save_screenshot(filename)
            logging.info(f"📸 Screenshot '{label}' saved at: {filename}")
            setattr(self, f"screenshot_{label}", filename)  # Dynamically track it
        except Exception as e:
            logging.warning(f"❌ Could not take screenshot for '{label}': {e}")

    def save_to_excel(self, info: dict):
        SCREENSHOT_LABEL = "📷 Screenshot"
        LOG_LABEL = "🧾 Log File"
        parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(parent_folder, "TestResults.xlsx")

        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results"
                ws.append([
                    "Test Name", "Order Number", "ID Number", "Description", "Status",
                    "Timestamp", "Duration", "Browser", "OS", "Full Name", "Email",
                    "Room Screenshot", "Payment Screenshot", "Confirmation Screenshot", "Log File"
                ])
            else:
                wb = load_workbook(filename)
                ws = wb.active

            status = "FAILED" if info.get("error") else "PASSED"

            row = [
                info.get("name", ""),
                info.get("order_number", ""),
                info.get("id_number", ""),
                info.get("description", ""),
                status,
                info.get("timestamp", ""),
                info.get("duration", ""),
                info.get("browser", ""),
                info.get("os", ""),
                info.get("full_name", ""),
                info.get("email", ""),
                getattr(self, "screenshot_room_selection", ""),
                getattr(self, "screenshot_payment_stage", ""),
                info.get("confirmation_screenshot" if info.get("status") == "PASSED" else "error_screenshot", ""),
                info.get("log", "")
            ]
            ws.append(row)
            row_num = ws.max_row

            # 🎨 Color row based on result
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row_fill = red_fill if status == "FAILED" else green_fill

            for col_idx in range(1, 16):  # Columns A to O (1-15)
                ws.cell(row=row_num, column=col_idx).fill = row_fill

            # 📎 Hyperlink screenshots with label
            for col_idx in [12, 13, 14]:
                screenshot_path = row[col_idx - 1]
                if screenshot_path and os.path.exists(screenshot_path):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = SCREENSHOT_LABEL
                    cell.hyperlink = f"file:///{screenshot_path.replace(os.sep, '/')}"
                    cell.font = Font(color="0000EE", underline="single")

            # 📎 Hyperlink log file with label
            log_path = info.get("log", "")
            if log_path and os.path.exists(log_path):
                cell = ws.cell(row=row_num, column=15)
                cell.value = LOG_LABEL
                cell.hyperlink = f"file:///{log_path.replace(os.sep, '/')}"
                cell.font = Font(color="0000EE", underline="single")

            # ↔️ Auto-fit column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.warning(f"Column resize error: {e}")
                ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(filename)
            logging.info(f"Excel file saved at: {filename}")

        except PermissionError as e:
            logging.warning(f"Excel file is open or locked: {e}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")

    def save_to_html(self, info: dict):
        html_dir = os.path.join(self.base_dir, "html_reports")
        os.makedirs(html_dir, exist_ok=True)

        dashboard_path = os.path.join(html_dir, "TestDashboard.html")

        def image_tag(path):
            if path and os.path.exists(path):
                rel_path = os.path.relpath(path, start=html_dir).replace(os.sep, '/')
                return f'<img src="{rel_path}" onclick="openModal(this.src)" style="max-height:120px; border:1px solid #ccc; margin:5px; cursor: pointer;" />'
            return "<em>No screenshot</em>"

        room_img = image_tag(getattr(self, 'screenshot_room_selection', ''))
        pay_img = image_tag(getattr(self, 'screenshot_payment_stage', ''))
        confirm_img = image_tag(
            info.get('error_screenshot') if info.get('status') == 'FAILED' else info.get('confirmation_screenshot'))

        test_type_class = info.get("test_type", "desktop").lower()

        # 💾 Convert log file to relative path
        log_file_rel = ""
        if info.get("log") and os.path.exists(info["log"]):
            log_file_rel = os.path.relpath(info["log"], start=html_dir).replace(os.sep, '/')

        html_entry = f"""
        <div class="test-block {test_type_class}">
            <h2>{info.get('name')} — <span class="{'fail' if info.get('status') == 'FAILED' else 'pass'}">{info.get('status')}</span></h2>
            <p><strong>Description:</strong> {info.get('description')}</p>
            <p><strong>Timestamp:</strong> {info.get('timestamp')} | <strong>Duration:</strong> {info.get('duration')}</p>
            <p><strong>Browser:</strong> {info.get('browser')} | <strong>OS:</strong> {info.get('os')}</p>
            <p><strong>Guest:</strong> {info.get('full_name')} | <strong>Email:</strong> {info.get('email')}</p>
            <p><strong>Order #:</strong> {info.get('order_number')} | <strong>ID:</strong> {info.get('id_number')}</p>
            <p><strong>Log File:</strong> <a href="{log_file_rel}" target="_blank">View Log</a></p>
            {'<p style="color:red;"><strong>Error:</strong> ' + info['error'] + '</p>' if info.get('error') else ''}
            <div class="grid">
                <div><h4>Room</h4>{room_img}</div>
                <div><h4>Payment</h4>{pay_img}</div>
                <div><h4>{'Failure Screenshot' if info.get('status') == 'FAILED' else 'Confirmation'}</h4>{confirm_img}</div>
            </div>
            <hr>
        </div>
        """

        if not os.path.exists(dashboard_path):
            html_start = """<!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Fattal Selenium Test Dashboard</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            .pass { color: green; font-weight: bold; }
            .fail { color: red; font-weight: bold; }
            .grid { display: flex; gap: 20px; margin-top: 10px; flex-wrap: wrap; }
            .grid div { flex: 1; min-width: 250px; }
            .test-block { margin-bottom: 40px; }
            hr { border: 1px dashed #ccc; }
            img { display: block; max-width: 100%; }
            img:hover {
                transform: scale(1.05);
                transition: transform 0.2s ease-in-out;
                box-shadow: 0 0 10px rgba(0,0,0,0.4);
            }
            .summary-bar {
                background-color: #f5f5f5;
                border: 1px solid #ccc;
                padding: 10px;
                margin-bottom: 20px;
            }
            .filters {
                margin-bottom: 20px;
            }
            .filters label {
                margin-right: 15px;
                cursor: pointer;
                font-weight: normal;
            }
            .filters input[type="checkbox"] {
                margin-right: 5px;
                transform: scale(1.1);
                vertical-align: middle;
            }
            .modal {
                display: none;
                position: fixed;
                z-index: 999;
                padding-top: 40px;
                left: 0; top: 0;
                width: 100%; height: 100%;
                overflow: auto;
                background-color: rgba(0,0,0,0.85);
            }
            .modal-content {
                margin: auto;
                display: block;
                max-width: 90vw;
                max-height: 80vh;
                object-fit: contain;
                border-radius: 8px;
            }
            .modal-content, .close {
                animation-name: zoom;
                animation-duration: 0.3s;
            }
            @keyframes zoom {
                from {transform: scale(0)} to {transform: scale(1)}
            }
            .close {
                position: absolute;
                top: 15px;
                right: 35px;
                color: #fff;
                font-size: 40px;
                font-weight: bold;
                cursor: pointer;
            }
            @media (max-width: 768px) {
                .modal-content {
                    max-width: 95%;
                    transform: none !important;
                }
                img:hover {
                    transform: none;
                }
            }
        </style>
        <script>
            function updateSummary() {
                let total = document.querySelectorAll('.test-block').length;
                let passed = document.querySelectorAll('.test-block .pass').length;
                let failed = document.querySelectorAll('.test-block .fail').length;
                document.getElementById('summary-total').textContent = total;
                document.getElementById('summary-passed').textContent = passed;
                document.getElementById('summary-failed').textContent = failed;
            }

            function applyFilters() {
                const showPassed = document.getElementById('filter-pass').checked;
                const showFailed = document.getElementById('filter-fail').checked;
                const showMobile = document.getElementById('filter-mobile').checked;
                const showDesktop = document.getElementById('filter-desktop').checked;

                document.querySelectorAll('.test-block').forEach(block => {
                    const isPass = block.querySelector('span').classList.contains('pass');
                    const isFail = block.querySelector('span').classList.contains('fail');
                    const isMobile = block.classList.contains('mobile');
                    const isDesktop = block.classList.contains('desktop');

                    let visible = false;
                    if ((showPassed && isPass) || (showFailed && isFail)) {
                        if ((showMobile && isMobile) || (showDesktop && isDesktop)) {
                            visible = true;
                        }
                    }

                    block.style.display = visible ? 'block' : 'none';
                });
            }

            function openModal(imgSrc) {
                const modal = document.getElementById("screenshotModal");
                const modalImg = document.getElementById("modalImage");
                modal.style.display = "block";
                modalImg.src = imgSrc;
            }
            function closeModal() {
                document.getElementById("screenshotModal").style.display = "none";
            }
            window.onload = function() {
                updateSummary();
                applyFilters();
            };
        </script>
    </head>
    <body>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h1 style="margin: 0;">🧪 Fattal Selenium Test Dashboard</h1>
            <img src="fattal_logo.png" alt="Fattal Logo" style="height: 150px;" />
        </div>
        <div class="summary-bar">
            <strong>Total Tests:</strong> <span id="summary-total">-</span> |
            <strong style="color:green;">Passed:</strong> <span id="summary-passed">-</span> |
            <strong style="color:red;">Failed:</strong> <span id="summary-failed">-</span>
        </div>
        <div class="filters">
            <label><input type="checkbox" id="filter-pass" checked onchange="applyFilters()">✅ Passed</label>
            <label><input type="checkbox" id="filter-fail" checked onchange="applyFilters()">❌ Failed</label>
            <label><input type="checkbox" id="filter-desktop" checked onchange="applyFilters()">💻 Desktop</label>
            <label><input type="checkbox" id="filter-mobile" checked onchange="applyFilters()">📱 Mobile</label>
        </div>
    """

            with open(dashboard_path, "w", encoding="utf-8") as f:
                f.write(html_start)

        with open(dashboard_path, "a", encoding="utf-8") as f:
            f.write(html_entry)

        with open(dashboard_path, "a", encoding="utf-8") as f:
            f.write("""
        <div id="screenshotModal" class="modal" onclick="closeModal()">
          <span class="close">&times;</span>
          <img class="modal-content" id="modalImage">
        </div>
    </body>
    </html>
    """)

    def take_confirmation_screenshot(self, test_method, status):
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"confirmation_{status}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
        )
        self.driver.save_screenshot(filename)
        logging.info(f"Confirmation screenshot saved at: {filename}")
        return filename

    def run(self, result=None):
        if result is None:
            result = self.defaultTestResult()
        self._test_result_for_teardown = result
        return super().run(result)

    def fill_payment_details(self):
        """
            Fills in credit card payment form fields from self.payment_card inside iframe context.
        """
        self.order_page.switch_to_payment_iframe()
        self.order_page.set_cardholder_name(self.payment_card["cardholder_name"])
        self.order_page.set_card_number(self.payment_card["card_number"])
        self.order_page.select_expiry_month(self.payment_card["expiry_month"])
        self.order_page.select_expiry_year(self.payment_card["expiry_year"])
        self.order_page.set_cvv(self.payment_card["cvv"])
        self.order_page.set_id_number_card(self.payment_card["id_number"])
        self.order_page.switch_to_default_content()

    def take_screenshot(self, test_method_name):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        screenshot_dir = os.path.join(self.base_dir, "Screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filename = os.path.join(
            screenshot_dir, f"{test_method_name}_{timestamp}.png"
        )
        self.driver.save_screenshot(filename)
        logging.error(f" Screenshot taken: {filename}")

    def retry_search_if_no_results(self, hotel_name, adults, children, infants):
        def no_results_displayed():
            try:
                msg = self.driver.find_element(By.ID, "search-page-no-search-results-title")
                return "לא נמצאו מלונות" in msg.text
            except:
                return False

        if no_results_displayed():
            logging.warning("No hotel-only results found. Retrying...")

            self.take_screenshot("hotel_search_no_results_initial")

            self.toolbar.logo_mainpage()

            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "main-input"))
            )

            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range()
            self.main_page.search_button()

            try:
                WebDriverWait(self.driver, 15).until_not(
                    EC.presence_of_element_located((By.ID, "search-page-no-search-results-title"))
                )
                logging.info("Hotel results appeared after retry.")
            except TimeoutException:
                self.take_screenshot("hotel_search_retry_failed")
                raise Exception("No hotel results even after retry.")
        else:
            logging.info("Hotel search results returned.")

    def complete_booking_post_flight(self):
        self.order_page.switch_to_default_content()
        self.order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")

        random_id = self.order_page.generate_israeli_id()
        self.entered_id_number = random_id  # 💾 For Excel logging

        logging.info(f"Generated Israeli ID: {random_id}")

        guest = self.default_guest
        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]
        self.entered_phone = guest["phone"]

        self.order_page.set_email(self.entered_email)
        self.order_page.set_phone(self.entered_phone)
        self.order_page.set_first_name(self.entered_first_name)
        self.order_page.set_last_name(self.entered_last_name)
        self.order_page.set_id_number(random_id)

        self.order_page.click_terms_approval_checkbox_js()
        sleep(10)
        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("בדיקת טסט נא לבטל")
        # for checkbox in [
        #     self.order_page.get_adjacent_rooms_checkbox(),
        #     self.order_page.get_high_floor_checkbox(),
        #     self.order_page.get_low_floor_checkbox()
        # ]:
        #     self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
        #     self.driver.execute_script("arguments[0].click();", checkbox)

        self.fill_payment_details()

        try:
            self.order_page.click_submit_button()
            logging.info(" Submit button clicked successfully.")
        except Exception as e:
            logging.error(f" Failed to click submit button: {e}")
            self.take_screenshot("submit_click_failure")
            raise

        try:
            WebDriverWait(self.driver, 25).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//h1[contains(text(), 'אישור') or contains(text(), 'תודה')]"))
            )
            logging.info(" Confirmation page loaded.")
        except:
            logging.warning(" Confirmation page not found yet. Continuing anyway.")

    def complete_booking_flow(self, hotel_name, adults, children, infants):
        random_id = self.order_page.generate_israeli_id()
        self.entered_id_number = random_id  # 💾 For Excel logging

        logging.info(f"Generated Israeli ID: {random_id}")

        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f" Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.take_stage_screenshot("room_selection")
        self.search_result.click_first_book_room()

        self.order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")

        guest = self.default_guest
        self.order_page.set_email(guest["email"])
        self.order_page.set_phone(guest["phone"])
        self.order_page.set_first_name(guest["first_name"])
        self.order_page.set_last_name(guest["last_name"])
        self.order_page.set_id_number(random_id)

        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]

        self.order_page.click_terms_approval_checkbox_js()
        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        sleep(10)
        self.order_page.click_terms_approval_checkbox_js()
        textarea.send_keys("בדיקת טסט נא לבטל")
        self.fill_payment_details()

        try:
            self.order_page.click_submit_button()
            logging.info(" Submit button clicked successfully.")
        except Exception as e:
            logging.error(f" Failed to click submit button: {e}")
            self.take_screenshot("submit_click_failure")
            raise

    def complete_booking_flow_club_checkbox(self, hotel_name, adults, children, infants):
        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f" Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.take_stage_screenshot("room_selection")

        self.search_result.click_first_book_room()

        self.order_page.switch_to_default_content()
        self.order_page.wait_until_personal_form_ready()
        self.order_page.club_checkbox()
        self.take_stage_screenshot("payment_stage")
        random_id = self.order_page.generate_israeli_id()
        self.entered_id_number = random_id  # 💾 For Excel logging

        logging.info(f"Generated Israeli ID: {random_id}")

        guest = self.default_guest

        self.order_page.set_email(guest["email"])
        self.order_page.set_phone(guest["phone"])
        self.order_page.set_first_name(guest["first_name"])
        self.order_page.set_last_name(guest["last_name"])
        self.order_page.set_id_number(random_id)

        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]

        self.order_page.click_terms_approval_checkbox_js()
        sleep(10)
        self.order_page.expand_special_requests_section()

        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("בדיקת טסט נא לבטל")

        # for checkbox in [
        #     self.order_page.get_adjacent_rooms_checkbox(),
        #     self.order_page.get_high_floor_checkbox(),
        #     self.order_page.get_low_floor_checkbox()
        # ]:
        #     self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
        #     self.driver.execute_script("arguments[0].click();", checkbox)

        self.fill_payment_details()

        try:
            self.order_page.click_submit_button()
            logging.info(" Submit button clicked successfully.")
        except Exception as e:
            logging.error(f" Failed to click submit button: {e}")
            self.take_screenshot("submit_click_failure")
            raise

    def test_desktop_anonymous_booking(self):
        self.save_for_cancellation = True
        self.soft_assert_errors = []

        hotel_name = self.default_hotel_name
        self.test_description = "בדיקת השלמת הזמנה מתשמש אנונימי"
        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.close_war_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0

        self.complete_booking_flow(hotel_name, adults, children, infants)

        # 👇 Add ID back into confirmation result
        self.confirm_and_assert_order()
        logging.info("✔️ Club login test finished with confirmed order.")
    def test_desktop_booking_anonymous_join_fattal_and_friends(self):
        self.save_for_cancellation = True
        self.soft_assert_errors = []

        self.test_description = "בידקת השלמת הזמנה משתמש אנונימי + הצטפרות למועדון"
        hotel_name = self.default_hotel_name

        logging.info(" Starting test: hotel search and booking flow")

        self.main_page.close_war_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()
        #self.main_page.select_random_date_range_two_months_ahead()

        adults, children, infants = 2, 0, 0

        self.complete_booking_flow_club_checkbox(hotel_name, adults, children, infants)
        self.confirm_and_assert_order()

    def test_desktop_booking_club_member_eilat_with_flight(self):
        self.save_for_cancellation = False
        self.soft_assert_errors = []
        guest = self.default_guest
        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]
        self.entered_phone = guest["phone"]

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר עם מועדון פעיל + טיסות"
        hotel_name = "אילת,ישראל"
        adults, children, infants = 2, 0, 0

        try:
            logging.info("Starting test for Eilat zone with flight")

            self.main_page.close_war_popup()
            try:
                self.toolbar.personal_zone()
                self.toolbar.click_footer_login_with_id_and_password()
                WebDriverWait(self.driver, 5).until(
                    EC.visibility_of(self.toolbar.user_id_input())
                ).send_keys("999318330")
                self.toolbar.user_password_input().send_keys("Aa123456")
                self.toolbar.login_button()
                self.main_page.close_post_login_popup()
                logging.info("Logged in to club account successfully.")
            except Exception as e:
                logging.warning(f"Login failed: {e}")
            self.main_page.close_war_popup()
            self.main_page.click_clear_button_hotel()
            self.main_page.set_city(hotel_name)
            self.main_page.select_next_month_date_range_eilat()

            self.main_page.set_room_occupants(adults, children, infants)
            self.main_page.select_flight_option_all_airports()
            self.main_page.search_button()

            self.retry_flight_search_if_no_results(hotel_name)

            self.search_result.click_book_room_button()
            self.search_result.wait_for_prices_to_load()
            self.search_result.click_first_show_prices()

            # 💡 Take room selection screenshot
            self.take_stage_screenshot("room_selection")

            self.search_result.click_first_book_room()

            # Optional flight selection
            try:
                self.flight_page.try_flight_options_by_time_of_day()
            except Exception as e:
                logging.warning(f"Flight selection failed (non-blocking): {e}")

            # Passenger form handling
            handled = self.flight_page.handle_passenger_form_if_flight_selection_skipped()
            if not handled:
                self.flight_page.wait_for_passenger_form()
                self.flight_page.fill_passenger_details_and_validate()
                self.flight_page.click_continue_button()

            # 💳 Payment screenshot is already inside complete_booking_post_flight
            self.complete_booking_post_flight()

            # ✅ Take confirmation screenshot manually here (just to be safe)
            self.take_stage_screenshot("confirmation_page")

        except Exception as e:
            logging.error(f"Test failed: {e}")
            try:
                self.take_screenshot("test_eilat_flight")
            except Exception:
                logging.warning("Could not capture screenshot due to earlier failure.")
            raise
        self.entered_email = self.default_guest["email"]
        self.confirm_and_assert_order()

    def test_desktop_booking_anonymous_region_eilat(self):
        self.save_for_cancellation = False
        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש אנונימי דרך אזור מלונות אילת"
        hotel_name = "אילת,ישראל"

        logging.info("Starting test: hotel search and booking flow")

        self.main_page.close_war_popup()
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range_eilat()

        #self.main_page.select_random_date_range_two_months_ahead()

        adults, children, infants = 2, 1, 0

        self.main_page.set_room_occupants(adults, children, infants)
        self.main_page.search_button()

        self.search_result.wait_for_rooms_to_load()
        self.search_result.click_book_room_button()
        self.search_result.wait_for_prices_to_load()
        self.search_result.click_first_show_prices()
        self.take_stage_screenshot("room_selection")

        self.search_result.click_first_book_room()

        self.complete_booking_post_flight()

        self.entered_email = self.default_guest["email"]
        self.entered_id_number = self.confirmation_result.get("id_number", "")
        self.confirm_and_assert_order()

    def test_desktop_booking_club_member(self):
        self.save_for_cancellation = True
        self.soft_assert_errors = []

        self.test_description = "בדיקת השלמת הזמנה משתמש מחובר חבר מועדון פעיל"
        hotel_name = self.default_hotel_name

        logging.info(" Starting test: hotel search and booking flow")

        guest = self.default_guest
        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]
        self.entered_phone = guest["phone"]

        self.main_page.close_war_popup()
        try:
            self.toolbar.personal_zone()
            self.toolbar.click_footer_login_with_id_and_password()
            WebDriverWait(self.driver, 5).until(
                EC.visibility_of(self.toolbar.user_id_input())
            ).send_keys("999318330")
            self.toolbar.user_password_input().send_keys("Aa123456")
            self.toolbar.login_button()
            self.main_page.close_post_login_popup()
            logging.info("Logged in to club account successfully.")
        except Exception as e:
            logging.warning(f"Login failed: {e}")
        self.main_page.click_clear_button_hotel()
        self.main_page.set_city(hotel_name)
        self.main_page.select_next_month_date_range()

        adults, children, infants = 2, 1, 0
        self.main_page.set_room_occupants(adults=adults, children=children, infants=infants)
        logging.info(f" Guests: {adults} adults, {children} children, {infants} infants")

        self.main_page.search_button()
        self.retry_search_if_no_results(hotel_name, adults, children, infants)

        self.search_result.click_first_show_prices()
        self.take_stage_screenshot("room_selection")
        self.search_result.click_first_book_room()

        self.order_page.wait_until_personal_form_ready()
        self.take_stage_screenshot("payment_stage")

        guest = self.default_guest
        self.order_page.set_email(guest["email"])
        self.order_page.set_phone(guest["phone"])
        self.order_page.set_first_name(guest["first_name"])
        self.order_page.set_last_name(guest["last_name"])

        random_id = self.order_page.generate_israeli_id()
        self.entered_id_number = random_id
        self.order_page.set_id_number(random_id)

        self.entered_email = guest["email"]
        self.entered_first_name = guest["first_name"]
        self.entered_last_name = guest["last_name"]

        self.order_page.click_terms_approval_checkbox_js()
        sleep(10)
        self.order_page.expand_special_requests_section()
        textarea = self.order_page.get_special_request_textarea()
        textarea.send_keys("בדיקת טסט נא לבטל")

        self.fill_payment_details()

        try:
            self.order_page.click_submit_button()
            logging.info(" Submit button clicked successfully.")
        except Exception as e:
            logging.error(f" Failed to click submit button: {e}")
            self.take_screenshot("submit_click_failure")
            raise

        self.confirmation_result = self.confirm_page.verify_confirmation_and_extract_order(self.entered_email)
        self.confirmation_result["id_number"] = self.entered_id_number

        self.confirm_and_assert_order()
        logging.info("✔️ Club login test finished with confirmed order.")

    @classmethod
    def tearDownClass(cls):
        try:
            runs_base_dir = os.path.join(os.path.dirname(__file__), "html_reports", "runs")
            cls.generate_dashboard_html(runs_base_dir)
            logging.info("✅ Dashboard generated once at the end of test suite.")
        except Exception as e:
            logging.warning(f"⚠️ Failed to generate dashboard in tearDownClass: {e}")
    def tearDown(self):
        if self.driver:
            try:
                # Post-test logging
                if hasattr(self, "_test_result_for_teardown"):
                    self.post_test_logging(self._test_result_for_teardown)
                else:
                    logging.warning("Test result object missing. Skipping logging.")

                # ⬇️ Try to save order for cancellation if flagged
                try:
                    confirmation = getattr(self, "confirmation_result", {})
                    order_number = confirmation.get("order_number")
                    hotel_id = HOTEL_NAME_TO_ID.get(self.default_hotel_name)
                    if getattr(self, "save_for_cancellation", False) and order_number and hotel_id:
                        save_order_for_cancellation(order_number, hotel_id)
                        logging.info(f"✅ Saved order {order_number} for cancellation.")
                    else:
                        logging.info("Skipping order save (either not flagged or missing data).")
                except Exception as e:
                    logging.warning(f"⚠️ Failed to write to cancellation JSON: {e}")

                # Attempt runtime tracking
                try:
                    duration = self.driver.execute_script("""
                        const [start] = window.performance.getEntriesByName('selenium-start');
                        return Date.now() - start.startTime;
                    """)
                    logging.info(f"🕒 Test runtime: {int(duration)}ms")
                except Exception as e:
                    logging.warning(f"⚠️ Could not get test runtime: {e}")

            except Exception as e:
                logging.warning(f"⚠️ Logging failed during tearDown: {e}")

            logging.info("⌛ Waiting 2 seconds before closing browser...")
            logging.info("🔚 Closing browser (tearDown).")
            try:
                self.driver.quit()
            except Exception as e:
                logging.warning(f"⚠️ Browser quit failed: {e}")

        # 🧾 Generate the HTML dashboard
        try:
            runs_base_dir = os.path.join(self.base_dir, "html_reports", "runs")
        except Exception as e:
            logging.warning(f"⚠️ Failed to generate dashboard: {e}")

    if __name__ == "__main__":
        import unittest
        unittest.main()











